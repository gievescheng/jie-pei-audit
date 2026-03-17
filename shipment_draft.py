from __future__ import annotations

import json
import re
import tempfile
from datetime import date, datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent.resolve()


def _read_json(name: str, fallback):
    path = BASE_DIR / name
    if not path.exists():
        return fallback
    return json.loads(path.read_text(encoding='utf-8'))


def _text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _parse_int(value):
    text = _text(value)
    if not text:
        return None
    match = re.search(r'(\d+)', text.replace(',', ''))
    return int(match.group(1)) if match else None


def _normalize_date(value) -> str:
    if not value:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _text(value)
    if not text:
        return ''

    if ' ' in text:
        text = text.split(' ', 1)[0]

    if re.fullmatch(r'\d{4}-\d{2}-\d{2}', text):
        return text
    if re.fullmatch(r'\d{4}/\d{2}/\d{2}', text):
        return text.replace('/', '-')
    if re.fullmatch(r'\d{3}[./]\d{2}[./]\d{2}', text):
        parts = re.split(r'[./]', text)
        year = int(parts[0]) + 1911
        return f'{year:04d}-{int(parts[1]):02d}-{int(parts[2]):02d}'
    return ''


def _display_date(iso_date: str) -> str:
    if not iso_date:
        return ''
    return iso_date.replace('-', '/')


def _product_defaults(customer_code: str, source_product: str) -> dict:
    combined = f'{customer_code} {source_product}'
    if any(token in combined for token in ('C001', '待洗玻璃', '12基板', '委外代工清洗')):
        return {
            'product_name_suggested': 'RECYCLE GLASS NEG ABC-1 (JEPE)',
            'spec_suggested': '12吋',
            'unit_suggested': '片',
        }
    return {
        'product_name_suggested': source_product,
        'spec_suggested': '',
        'unit_suggested': '片',
    }


def _ensure_entry(catalog: dict, order_no: str) -> dict:
    entry = catalog.setdefault(order_no, {
        'order_no': order_no,
        'customer_code': '',
        'source_product': '',
        'order_qty': None,
        'plan_qty': None,
        'due_date': '',
        'ship_date_suggested': '',
        'department_suggested': '資材課',
        'requester_suggested': '',
        'remark_suggested': '',
        'lots': [],
        '_lot_map': {},
    })
    return entry


def _upsert_lot(entry: dict, lot: str, payload: dict) -> None:
    if not lot:
        return
    if lot in {'全部', 'ALL', 'All'}:
        return
    bucket = entry['_lot_map']
    current = bucket.get(lot)
    if current is None:
        current = {
            'lot': lot,
            'date': '',
            'plan_qty': None,
            'actual_qty': None,
            'good': None,
            'bad': None,
            'operator': '',
        }
        bucket[lot] = current
        entry['lots'].append(current)

    for key, value in payload.items():
        if value in ('', None):
            continue
        if key in ('plan_qty', 'actual_qty', 'good', 'bad'):
            current[key] = value
        else:
            current[key] = value


def _load_inspect_111(catalog: dict) -> None:
    data = _read_json('inspect_111.json', {})
    for rows in data.values():
        if not isinstance(rows, list) or len(rows) < 5:
            continue
        row4 = rows[3] if len(rows) > 3 else []
        row5 = rows[4] if len(rows) > 4 else []
        order_no = _text(row5[4] if len(row5) > 4 else '')
        if not order_no:
            order_no = _text(row4[4] if len(row4) > 4 else '')
        if not order_no:
            continue

        entry = _ensure_entry(catalog, order_no)
        entry['customer_code'] = entry['customer_code'] or _text(row4[2] if len(row4) > 2 else '')
        entry['source_product'] = entry['source_product'] or _text(row4[3] if len(row4) > 3 else '')
        entry['order_qty'] = entry['order_qty'] or _parse_int(row4[5] if len(row4) > 5 else '')
        entry['due_date'] = entry['due_date'] or _normalize_date(row5[6] if len(row5) > 6 else '')


def _load_verify2(catalog: dict) -> None:
    data = _read_json('verify2.json', {})
    for value in data.values():
        if not isinstance(value, list):
            continue
        for row in value:
            order_no = _text(row.get('order'))
            if not order_no:
                continue
            entry = _ensure_entry(catalog, order_no)
            entry['source_product'] = entry['source_product'] or _text(row.get('product'))
            plan_qty = _parse_int(row.get('plan_qty'))
            if plan_qty:
                entry['plan_qty'] = max(entry['plan_qty'] or 0, plan_qty)
            plan_date = _normalize_date(row.get('plan_date'))
            if plan_date and not entry['ship_date_suggested']:
                entry['ship_date_suggested'] = plan_date
            _upsert_lot(entry, _text(row.get('fosb')), {
                'date': _normalize_date(row.get('date')),
                'plan_qty': plan_qty,
                'actual_qty': _parse_int(row.get('actual_qty')),
                'good': _parse_int(row.get('good')),
                'bad': _parse_int(row.get('bad')),
                'operator': _text(row.get('operator')),
            })


def _load_inspect_115(catalog: dict) -> None:
    data = _read_json('inspect_115.json', {})
    for payload in data.values():
        if not isinstance(payload, dict):
            continue
        rows = payload.get('data') or []
        if len(rows) < 5:
            continue

        header = rows[2]
        order_no = _text(header[9] if len(header) > 9 else '')
        if not order_no:
            continue
        create_date = _normalize_date(header[11] if len(header) > 11 else '')
        entry = _ensure_entry(catalog, order_no)
        if create_date and not entry['ship_date_suggested']:
            entry['ship_date_suggested'] = create_date

        current_product = ''
        for row in rows[4:]:
            product = _text(row[2] if len(row) > 2 else '')
            if product:
                current_product = product
                entry['source_product'] = entry['source_product'] or product
            lot = _text(row[3] if len(row) > 3 else '')
            if not lot:
                continue
            _upsert_lot(entry, lot, {
                'date': create_date,
                'actual_qty': _parse_int(row[4] if len(row) > 4 else ''),
                'good': _parse_int(row[6] if len(row) > 6 else ''),
                'bad': _parse_int(row[7] if len(row) > 7 else ''),
                'operator': _text(row[10] if len(row) > 10 else ''),
            })


def _finalize_catalog(catalog: dict) -> list[dict]:
    orders = []
    for order_no, entry in catalog.items():
        defaults = _product_defaults(entry['customer_code'], entry['source_product'])
        quantity = entry['order_qty'] or entry['plan_qty']
        if quantity is None and entry['lots']:
            quantity = sum(item.get('good') or item.get('actual_qty') or 0 for item in entry['lots'])

        lots = sorted(
            entry['lots'],
            key=lambda item: ((item.get('date') or ''), item.get('lot') or ''),
        )
        batch_display = order_no
        if 0 < len(lots) <= 4:
            batch_display = f"{order_no}/" + ','.join(item['lot'] for item in lots)

        item = {
            'order_no': order_no,
            'customer_code': entry['customer_code'],
            'source_product': entry['source_product'],
            'product_name_suggested': defaults['product_name_suggested'],
            'spec_suggested': defaults['spec_suggested'],
            'quantity_suggested': quantity or '',
            'unit_suggested': defaults['unit_suggested'],
            'ship_date_suggested': entry['ship_date_suggested'] or entry['due_date'],
            'due_date': entry['due_date'],
            'department_suggested': entry['department_suggested'],
            'requester_suggested': entry['requester_suggested'],
            'remark_suggested': entry['remark_suggested'],
            'batch_display_suggested': batch_display,
            'lots': lots,
        }
        orders.append(item)

    return sorted(orders, key=lambda item: item['order_no'])


def get_order_catalog() -> list[dict]:
    catalog: dict[str, dict] = {}
    _load_inspect_111(catalog)
    _load_verify2(catalog)
    _load_inspect_115(catalog)
    return _finalize_catalog(catalog)


def _template_path() -> Path:
    candidates = [
        path for path in BASE_DIR.rglob('*.xlsx')
        if path.name == '14.3出貨單.xlsx' and path.parent.name == '表單'
    ]
    if not candidates:
        raise FileNotFoundError('14.3出貨單.xlsx template not found.')
    return sorted(candidates, key=lambda item: len(item.parts))[0]


def _build_batch_display(order_no: str, selected_lots: list[str], fallback: str) -> str:
    lots = [lot.strip() for lot in selected_lots if lot and lot.strip()]
    if lots:
        return f"{order_no}/" + ','.join(lots)
    if fallback:
        return fallback.strip()
    return order_no


def build_shipment_draft(payload: dict) -> tuple[str, str]:
    catalog = {item['order_no']: item for item in get_order_catalog()}
    order_no = _text(payload.get('order_no'))
    if not order_no:
        raise ValueError('order_no is required.')

    defaults = catalog.get(order_no, {
        'order_no': order_no,
        'department_suggested': '資材課',
        'requester_suggested': '',
        'ship_date_suggested': date.today().isoformat(),
        'product_name_suggested': '',
        'spec_suggested': '',
        'quantity_suggested': '',
        'unit_suggested': '片',
        'remark_suggested': '',
        'batch_display_suggested': order_no,
    })

    selected_lots = payload.get('selected_lots') or []
    if not isinstance(selected_lots, list):
        selected_lots = []

    draft_date = _normalize_date(payload.get('date')) or defaults.get('ship_date_suggested') or date.today().isoformat()
    department = _text(payload.get('department')) or defaults.get('department_suggested') or '資材課'
    requester = _text(payload.get('requester')) or defaults.get('requester_suggested') or ''
    product_name = _text(payload.get('product_name')) or defaults.get('product_name_suggested') or defaults.get('source_product') or ''
    spec = _text(payload.get('spec')) or defaults.get('spec_suggested') or ''
    unit = _text(payload.get('unit')) or defaults.get('unit_suggested') or '片'
    remark = _text(payload.get('remark')) or defaults.get('remark_suggested') or ''
    batch_display = _build_batch_display(
        order_no,
        selected_lots,
        _text(payload.get('batch_display')) or defaults.get('batch_display_suggested') or order_no,
    )

    quantity = payload.get('quantity')
    quantity_int = _parse_int(quantity)
    if quantity_int is None:
        quantity_int = _parse_int(defaults.get('quantity_suggested'))

    template_path = _template_path()
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, dir=tempfile.gettempdir())
    tmp.close()
    copy2(template_path, tmp.name)

    wb = load_workbook(tmp.name)
    ws = wb.active
    ws['B3'] = department
    ws['D3'] = requester
    ws['E3'] = f"日期：{_display_date(draft_date)}"
    ws['A5'] = 1
    ws['B5'] = batch_display
    ws['C5'] = product_name
    ws['D5'] = spec
    ws['E5'] = quantity_int if quantity_int is not None else None
    ws['F5'] = unit
    ws['G5'] = remark or None
    wb.save(tmp.name)

    download_name = f'{order_no}_出貨單草稿.xlsx'
    return tmp.name, download_name
