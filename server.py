from __future__ import annotations

import hashlib
import json
import os
import secrets
import tempfile
import time
import traceback
from functools import wraps
from pathlib import Path
from urllib.parse import urlencode

import jwt as pyjwt
import requests
from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    redirect,
    request,
    send_file,
    send_from_directory,
    session,
)
import ops_data
from runtime_paths import (
    GOOGLE_CONFIG_PATH,
    GOOGLE_TOKEN_PATH,
    get_or_create_flask_secret,
    migrate_legacy_private_files,
    public_root_contains_private_files,
)

BASE_DIR = Path(__file__).parent.resolve()
GOOGLE_AUTH_URL = 'https://accounts.google.com/o/oauth2/v2/auth'
GOOGLE_TOKEN_URL = 'https://oauth2.googleapis.com/token'
GOOGLE_USERINFO_URL = 'https://www.googleapis.com/oauth2/v2/userinfo'
GOOGLE_CALENDAR_EVENTS_URL = 'https://www.googleapis.com/calendar/v3/calendars/primary/events'
GOOGLE_SCOPES = [
    'https://www.googleapis.com/auth/calendar.events',
    'https://www.googleapis.com/auth/userinfo.email',
    'https://www.googleapis.com/auth/userinfo.profile',
]
PUBLIC_STATIC_DIRS = {
    'vendor': BASE_DIR / 'vendor',
    'public': BASE_DIR / 'public',
}
PUBLIC_STATIC_FILES = {
    'index.html': BASE_DIR / 'index.html',
}

app = Flask(__name__, static_folder=None)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024
migrate_legacy_private_files()
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or get_or_create_flask_secret()


def read_json(path: Path, fallback):
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return fallback


def write_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')


def delete_file(path: Path) -> None:
    if path.exists():
        path.unlink()


def load_google_config() -> dict:
    return read_json(GOOGLE_CONFIG_PATH, {})


def save_google_config(client_id: str, client_secret: str) -> dict:
    data = {
        'client_id': client_id,
        'client_secret': client_secret,
        'updated_at': int(time.time()),
    }
    write_json(GOOGLE_CONFIG_PATH, data)
    return data


def clear_google_config() -> None:
    delete_file(GOOGLE_CONFIG_PATH)


def load_google_tokens() -> dict:
    return read_json(GOOGLE_TOKEN_PATH, {})


def save_google_tokens(tokens: dict) -> dict:
    tokens = dict(tokens)
    tokens['updated_at'] = int(time.time())
    write_json(GOOGLE_TOKEN_PATH, tokens)
    return tokens


def clear_google_tokens() -> None:
    delete_file(GOOGLE_TOKEN_PATH)


def google_redirect_uri() -> str:
    return request.host_url.rstrip('/') + '/api/google-calendar/oauth/callback'


def google_configured(config: dict | None = None) -> bool:
    config = config or load_google_config()
    return bool(config.get('client_id') and config.get('client_secret'))


def enrich_google_identity(tokens: dict) -> dict:
    access_token = tokens.get('access_token')
    if not access_token:
        return tokens
    try:
        response = requests.get(
            GOOGLE_USERINFO_URL,
            headers={'Authorization': f'Bearer {access_token}'},
            timeout=20,
        )
        if response.ok:
            userinfo = response.json()
            tokens['email'] = userinfo.get('email', tokens.get('email', ''))
            tokens['name'] = userinfo.get('name', tokens.get('name', ''))
    except Exception:
        pass
    return tokens


def exchange_google_code(code: str, config: dict) -> dict:
    response = requests.post(
        GOOGLE_TOKEN_URL,
        data={
            'code': code,
            'client_id': config['client_id'],
            'client_secret': config['client_secret'],
            'redirect_uri': google_redirect_uri(),
            'grant_type': 'authorization_code',
        },
        timeout=20,
    )
    if not response.ok:
        raise RuntimeError(google_error_text(response))
    tokens = response.json()
    tokens['expires_at'] = int(time.time()) + int(tokens.get('expires_in', 3600)) - 60
    return enrich_google_identity(tokens)


def refresh_google_tokens(tokens: dict, config: dict) -> dict:
    expires_at = int(tokens.get('expires_at', 0) or 0)
    if expires_at > int(time.time()) + 30:
        return tokens

    refresh_token = tokens.get('refresh_token')
    if not refresh_token:
        raise RuntimeError('Google Calendar authorization expired. Please reconnect Google Calendar.')

    response = requests.post(
        GOOGLE_TOKEN_URL,
        data={
            'client_id': config['client_id'],
            'client_secret': config['client_secret'],
            'refresh_token': refresh_token,
            'grant_type': 'refresh_token',
        },
        timeout=20,
    )
    if not response.ok:
        raise RuntimeError(google_error_text(response))

    refreshed = response.json()
    merged = dict(tokens)
    merged.update(refreshed)
    merged['refresh_token'] = refresh_token
    merged['expires_at'] = int(time.time()) + int(refreshed.get('expires_in', 3600)) - 60
    return save_google_tokens(enrich_google_identity(merged))


def require_google_access_token() -> tuple[str, dict]:
    config = load_google_config()
    if not google_configured(config):
        raise RuntimeError('Google Calendar is not configured. Fill in Client ID and Client Secret first.')

    tokens = load_google_tokens()
    if not tokens.get('access_token'):
        raise RuntimeError('Google Calendar is not connected. Complete Google authorization first.')

    tokens = refresh_google_tokens(tokens, config)
    return tokens['access_token'], tokens


def google_error_text(response: requests.Response) -> str:
    try:
        payload = response.json()
        if isinstance(payload.get('error'), dict):
            return payload['error'].get('message', json.dumps(payload, ensure_ascii=False))
        return payload.get('error_description') or payload.get('error') or json.dumps(payload, ensure_ascii=False)
    except Exception:
        return response.text.strip() or f'HTTP {response.status_code}'


def json_error(message: str, status_code: int = 400):
    return jsonify({'error': message}), status_code


def json_body() -> dict:
    return request.get_json(force=True) or {}


def save_ops_records(kind: str):
    body = json_body()
    records = body.get('records')
    replace_source_file = str(body.get('replace_source_file') or '').strip()
    if records is None:
        record = body.get('record')
        records = record if isinstance(record, list) else [record or body]
    if not isinstance(records, list) or not records:
        return json_error('No records provided.')
    items, saved = ops_data.upsert_records(
        kind,
        [record for record in records if isinstance(record, dict)],
        replace_source_file=replace_source_file,
    )
    payload = {'items': items, 'saved': saved}
    if kind == 'environment':
        payload['summary'] = ops_data.summarize_environment(items)
    return jsonify(payload)


def serve_managed_file(stored_path: str, as_attachment: bool):
    file_path = ops_data.get_serving_path(stored_path)
    if file_path is None:
        abort(404)
    return send_file(file_path, as_attachment=as_attachment, download_name=file_path.name)


def google_status_payload() -> dict:
    config = load_google_config()
    tokens = load_google_tokens()
    return {
        'configured': google_configured(config),
        'connected': bool(tokens.get('access_token')),
        'email': tokens.get('email', ''),
        'name': tokens.get('name', ''),
        'redirect_uri': google_redirect_uri(),
        'has_refresh_token': bool(tokens.get('refresh_token')),
        'expires_at': tokens.get('expires_at'),
    }


def build_event_payload(item: dict) -> dict:
    date_value = str(item.get('date', '')).strip()
    if len(date_value) != 10:
        raise ValueError('Each alert item must include a YYYY-MM-DD date.')

    title = str(item.get('title') or 'Audit reminder').strip()
    module = str(item.get('module') or '').strip()
    summary = str(item.get('summary') or '').strip()
    owner = str(item.get('owner') or '').strip()

    details = [
        f'Module: {module}' if module else '',
        f'Summary: {summary}' if summary else '',
        f'Owner: {owner}' if owner else '',
    ]
    return {
        'summary': title,
        'description': '\n'.join(line for line in details if line),
        'start': {'date': date_value},
        'end': {'date': add_days(date_value, 1)},
    }


def add_days(date_str: str, days: int) -> str:
    from datetime import date, timedelta

    base = date.fromisoformat(date_str)
    return (base + timedelta(days=days)).isoformat()


def resolve_public_static_file(filename: str) -> Path | None:
    cleaned = filename.strip().lstrip('/').replace('\\', '/')
    if not cleaned or cleaned.startswith('.'):
        return None
    if cleaned in PUBLIC_STATIC_FILES:
        path = PUBLIC_STATIC_FILES[cleaned]
        return path if path.exists() and path.is_file() else None

    for prefix, root in PUBLIC_STATIC_DIRS.items():
        if cleaned == prefix or not cleaned.startswith(prefix + '/'):
            continue
        relative = cleaned[len(prefix) + 1 :]
        candidate = (root / relative).resolve()
        try:
            candidate.relative_to(root.resolve())
        except ValueError:
            return None
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def startup_security_warnings(host: str) -> list[str]:
    warnings = []
    public_private_files = public_root_contains_private_files()
    if public_private_files:
        warnings.append('Private config files are still present in the public project root.')
    if host == '0.0.0.0':
        warnings.append('Server is bound to 0.0.0.0. Other machines on the network may reach this site.')
    if not os.environ.get('FLASK_SECRET_KEY'):
        warnings.append('FLASK_SECRET_KEY is not provided. Using a generated local secret file instead.')
    return warnings


@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')


@app.route('/<path:filename>')
def static_files(filename):
    filepath = resolve_public_static_file(filename)
    if filepath is not None:
        return send_from_directory(filepath.parent, filepath.name)
    abort(404)


@app.route('/api/files/view')
def api_view_file():
    return serve_managed_file(request.args.get('path', ''), as_attachment=False)


@app.route('/api/files/download')
def api_download_file():
    return serve_managed_file(request.args.get('path', ''), as_attachment=True)


@app.route('/api/files/preview-text')
def api_preview_text_file():
    html_payload = ops_data.build_text_preview_html(request.args.get('path', ''))
    if html_payload is None:
        abort(404)
    return Response(html_payload, mimetype='text/html')


@app.route('/api/nonconformances', methods=['GET'])
def api_nonconformances_list():
    return jsonify({'items': ops_data.load_records('nonconformance')})


@app.route('/api/nonconformances', methods=['POST'])
def api_nonconformances_save():
    try:
        return save_ops_records('nonconformance')
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/nonconformances/<record_id>', methods=['DELETE'])
def api_nonconformances_delete(record_id):
    try:
        items, deleted = ops_data.delete_record('nonconformance', record_id)
        if not deleted:
            return json_error('Record not found.', 404)
        return jsonify({'items': items, 'deleted_id': record_id})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/nonconformances/import', methods=['POST'])
def api_nonconformances_import():
    try:
        upload = request.files.get('file')
        if upload is None or not upload.filename:
            return json_error('Upload file is required.')
        return jsonify(ops_data.parse_import('nonconformance', upload))
    except ValueError as exc:
        return json_error(str(exc), 400)
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/audit-plans', methods=['GET'])
def api_audit_plans_list():
    return jsonify({'items': ops_data.load_records('auditplan')})


@app.route('/api/audit-plans', methods=['POST'])
def api_audit_plans_save():
    try:
        return save_ops_records('auditplan')
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/audit-plans/<record_id>', methods=['DELETE'])
def api_audit_plans_delete(record_id):
    try:
        items, deleted = ops_data.delete_record('auditplan', record_id)
        if not deleted:
            return json_error('Record not found.', 404)
        return jsonify({'items': items, 'deleted_id': record_id})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/audit-plans/import', methods=['POST'])
def api_audit_plans_import():
    try:
        upload = request.files.get('file')
        if upload is None or not upload.filename:
            return json_error('Upload file is required.')
        return jsonify(ops_data.parse_import('auditplan', upload))
    except ValueError as exc:
        return json_error(str(exc), 400)
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/audit-plans/<record_id>/attachments', methods=['GET'])
def api_audit_plan_attachments(record_id):
    try:
        return jsonify({'attachments': ops_data.list_auditplan_attachments(record_id)})
    except KeyError:
        return json_error('Record not found.', 404)
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/environment-records', methods=['GET'])
def api_environment_records_list():
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    items = ops_data.filter_environment_records(start, end)
    return jsonify({'items': items, 'summary': ops_data.summarize_environment(items)})


@app.route('/api/environment-records', methods=['POST'])
def api_environment_records_save():
    try:
        return save_ops_records('environment')
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/environment-records/<record_id>', methods=['DELETE'])
def api_environment_records_delete(record_id):
    try:
        items, deleted = ops_data.delete_record('environment', record_id)
        if not deleted:
            return json_error('Record not found.', 404)
        return jsonify({'items': items, 'summary': ops_data.summarize_environment(items), 'deleted_id': record_id})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/environment-records/import', methods=['POST'])
def api_environment_records_import():
    try:
        upload = request.files.get('file')
        if upload is None or not upload.filename:
            return json_error('Upload file is required.')
        return jsonify(ops_data.parse_import('environment', upload))
    except ValueError as exc:
        return json_error(str(exc), 400)
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/environment-records/delete-range', methods=['POST'])
def api_environment_records_delete_range():
    try:
        body = json_body()
        start = body.get('start', '')
        end = body.get('end', '')
        items, removed = ops_data.delete_environment_range(start, end)
        return jsonify({'items': items, 'summary': ops_data.summarize_environment(items), 'removed_count': removed, 'deleted': removed})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/production-records/read-existing', methods=['GET'])
def api_production_records_read_existing():
    try:
        import record_imports

        records, source_file = record_imports.load_existing_production_records()
        return jsonify({'records': records, 'source_file': source_file})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/production-records/import', methods=['POST'])
def api_production_records_import():
    temp_path = None
    try:
        import record_imports

        uploaded = request.files.get('file')
        if not uploaded or not uploaded.filename:
            return json_error('請先選擇生產日報 Excel 檔案。', 400)

        suffix = Path(uploaded.filename).suffix or '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            uploaded.save(tmp.name)
            temp_path = Path(tmp.name)

        records, _ = record_imports.load_uploaded_production_records(temp_path)
        if not records:
            return json_error('這份檔案沒有讀到可辨識的生產日報資料，請確認是否為正式生產日報格式。', 400)
        return jsonify({'records': records, 'source_file': uploaded.filename})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


@app.route('/api/quality-records/read-existing', methods=['GET'])
def api_quality_records_read_existing():
    try:
        import record_imports

        records, source_file = record_imports.load_existing_quality_records()
        return jsonify({'records': records, 'source_file': source_file})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)


@app.route('/api/quality-records/import', methods=['POST'])
def api_quality_records_import():
    temp_path = None
    try:
        import record_imports

        uploaded = request.files.get('file')
        if not uploaded or not uploaded.filename:
            return json_error('請先選擇品質檢驗 Excel 檔案。', 400)

        suffix = Path(uploaded.filename).suffix or '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            uploaded.save(tmp.name)
            temp_path = Path(tmp.name)

        records, _ = record_imports.load_uploaded_quality_records(temp_path)
        if not records:
            return json_error('這份檔案沒有讀到可辨識的品質檢驗資料，請確認是否為正式品質記錄格式。', 400)
        return jsonify({'records': records, 'source_file': uploaded.filename})
    except Exception as exc:
        traceback.print_exc()
        return json_error(str(exc), 500)
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


@app.route('/api/generate', methods=['POST'])
def api_generate():
    try:
        import generate_record

        body = request.get_json(force=True)
        rec_type = body.get('type', '')
        data = body.get('data', [])

        if rec_type not in ('env', 'production', 'quality'):
            return jsonify({'error': f'Unsupported type: {rec_type}'}), 400

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, dir=tempfile.gettempdir())
        tmp.close()
        out_path = generate_record.run(rec_type, data, tmp.name)

        filenames = {
            'env': '環境監測記錄.xlsx',
            'production': '生產記錄.xlsx',
            'quality': '品質檢驗記錄.xlsx',
        }
        return send_file(
            out_path,
            as_attachment=True,
            download_name=filenames[rec_type],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/record-engine/catalog', methods=['GET'])
def api_record_engine_catalog():
    try:
        import record_engine

        return jsonify({'templates': record_engine.get_catalog()})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/record-engine/suggest', methods=['POST'])
def api_record_engine_suggest():
    try:
        import record_engine

        body = request.get_json(force=True) or {}
        prompt = body.get('prompt', '')
        context = body.get('context') or {}
        return jsonify({'templates': record_engine.suggest_templates(prompt, context)})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/record-engine/precheck', methods=['POST'])
def api_record_engine_precheck():
    try:
        import record_engine

        body = request.get_json(force=True) or {}
        return jsonify({'result': record_engine.precheck_template(body)})
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/record-engine/generate', methods=['POST'])
def api_record_engine_generate():
    try:
        import record_engine

        body = request.get_json(force=True) or {}
        out_path, download_name, mimetype = record_engine.generate_template(body)
        return send_file(
            str(out_path),
            as_attachment=True,
            download_name=download_name,
            mimetype=mimetype,
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/notion', methods=['POST'])
def api_notion():
    try:
        body = request.get_json(force=True)
        token = body.get('token', '').strip()
        db_id = body.get('db_id', '').strip()
        properties = body.get('properties', {})

        if not token or not db_id:
            return jsonify({'error': 'token and db_id are required.'}), 400

        response = requests.post(
            'https://api.notion.com/v1/pages',
            json={'parent': {'database_id': db_id}, 'properties': properties},
            headers={
                'Authorization': f'Bearer {token}',
                'Notion-Version': '2022-06-28',
                'Content-Type': 'application/json',
            },
            timeout=20,
        )
        return jsonify(response.json()), response.status_code
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/shipment-draft/catalog', methods=['GET'])
def shipment_draft_catalog():
    try:
        import shipment_draft

        return jsonify({'orders': shipment_draft.get_order_catalog()})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/shipment-draft/generate', methods=['POST'])
def shipment_draft_generate():
    try:
        import shipment_draft

        body = request.get_json(force=True) or {}
        out_path, download_name = shipment_draft.build_shipment_draft(body)
        return send_file(
            out_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/google-calendar/status', methods=['GET'])
def google_calendar_status():
    return jsonify(google_status_payload())


@app.route('/api/google-calendar/config', methods=['POST'])
def google_calendar_config():
    body = request.get_json(force=True) or {}
    if body.get('clear'):
        clear_google_config()
        clear_google_tokens()
        return jsonify(google_status_payload())

    client_id = body.get('client_id', '').strip()
    client_secret = body.get('client_secret', '').strip()
    if not client_id or not client_secret:
        return jsonify({'error': 'client_id and client_secret are required.'}), 400

    save_google_config(client_id, client_secret)
    clear_google_tokens()
    return jsonify(google_status_payload())


@app.route('/api/google-calendar/auth/start', methods=['GET'])
def google_calendar_auth_start():
    config = load_google_config()
    if not google_configured(config):
        return jsonify({'error': 'Google Calendar is not configured yet.'}), 400

    state = secrets.token_urlsafe(24)
    session['google_oauth_state'] = state
    auth_url = GOOGLE_AUTH_URL + '?' + urlencode({
        'client_id': config['client_id'],
        'redirect_uri': google_redirect_uri(),
        'response_type': 'code',
        'scope': ' '.join(GOOGLE_SCOPES),
        'access_type': 'offline',
        'prompt': 'consent',
        'include_granted_scopes': 'true',
        'state': state,
    })
    return jsonify({'auth_url': auth_url, 'redirect_uri': google_redirect_uri()})


@app.route('/api/google-calendar/oauth/callback', methods=['GET'])
def google_calendar_oauth_callback():
    expected_state = session.get('google_oauth_state')
    incoming_state = request.args.get('state', '')
    if not expected_state or expected_state != incoming_state:
        return redirect('/?google=error&reason=state')

    if request.args.get('error'):
        return redirect('/?google=error&reason=' + request.args.get('error', 'oauth'))

    code = request.args.get('code', '')
    if not code:
        return redirect('/?google=error&reason=missing_code')

    try:
        config = load_google_config()
        tokens = exchange_google_code(code, config)
        save_google_tokens(tokens)
        session.pop('google_oauth_state', None)
        return redirect('/?google=connected')
    except Exception:
        traceback.print_exc()
        return redirect('/?google=error&reason=token_exchange')


@app.route('/api/google-calendar/events', methods=['POST'])
def google_calendar_events():
    try:
        body = request.get_json(force=True) or {}
        items = body.get('items')
        if items is None:
            items = [body]
        if not isinstance(items, list) or not items:
            return jsonify({'error': 'items must be a non-empty array.'}), 400

        access_token, tokens = require_google_access_token()
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json',
        }

        results = []
        success_count = 0
        for item in items:
            try:
                payload = build_event_payload(item)
                response = requests.post(GOOGLE_CALENDAR_EVENTS_URL, json=payload, headers=headers, timeout=20)
                if not response.ok:
                    raise RuntimeError(google_error_text(response))
                event = response.json()
                results.append({
                    'title': item.get('title', ''),
                    'success': True,
                    'event_id': event.get('id', ''),
                    'html_link': event.get('htmlLink', ''),
                })
                success_count += 1
            except Exception as exc:
                results.append({
                    'title': item.get('title', ''),
                    'success': False,
                    'error': str(exc),
                })

        failed_count = len(results) - success_count
        status_code = 200 if failed_count == 0 else 207
        return jsonify({
            'email': tokens.get('email', ''),
            'success_count': success_count,
            'failed_count': failed_count,
            'results': results,
        }), status_code
    except RuntimeError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'error': str(exc)}), 500


@app.route('/api/google-calendar/logout', methods=['POST'])
def google_calendar_logout():
    clear_google_tokens()
    return jsonify(google_status_payload())


# ─── SPC 統計製程管制 API ────────────────────────────────────────────────────

@app.route('/api/spc/analyze', methods=['POST'])
def api_spc_analyze():
    """
    接收 CSV 晶圓量測資料，呼叫 spc_engine 計算 I-MR 管制圖與 Cpk。
    Body（JSON）：
      {
        "batch_id": "LOT-20260301-A",
        "thickness": [702.1, 701.3, ...],
        "ttv":       [0.28, 0.19, ...],
        "particle_lots": [{"lot_id":"L001","n":600,"defects":3}, ...],  // 選填
        "spec": {
          "thickness_usl": 705.0, "thickness_lsl": 695.0,
          "ttv_usl": 2.0,         "ttv_lsl": 0.0
        }
      }
    或直接上傳 CSV 檔（multipart/form-data，欄位名 csv_file）。
    """
    try:
        from spc_engine import run_all_charts
    except ImportError:
        return json_error("spc_engine 模組未安裝，請確認 numpy/scipy 已安裝", 500)

    # ── CSV 上傳模式 ──────────────────────────────────────────
    raw_rows = []   # 原始量測列（含 Wafer ID），供 FOSB 報表使用

    if request.files.get("csv_file"):
        import csv, io
        f = request.files["csv_file"]
        spec_raw = request.form.get("spec", "{}")
        batch_id = request.form.get("batch_id", f.filename.rsplit(".", 1)[0])
        try:
            spec = json.loads(spec_raw)
        except Exception:
            spec = {}

        # 一次讀取所有內容，避免 stream 被消耗
        content = f.read().decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(content))
        csv_rows = list(reader)

        thickness, ttv, errors = [], [], []
        for i, row in enumerate(csv_rows, 1):
            wid   = (row.get("Wafer ID") or row.get("wafer_id") or row.get("slot") or "").strip()
            t_raw = row.get("Thickness Avg(um)", row.get("thickness", "") or "").strip()
            v_raw = row.get("TTV(um)",            row.get("ttv",       "") or "").strip()
            p_raw = row.get("Particle Count",     row.get("particle",  "") or "").strip()
            t_val = v_val = p_val = None
            try:
                if t_raw and t_raw not in ("N/A", "--"):
                    t_val = float(t_raw)
                    thickness.append(t_val)
            except ValueError as e:
                errors.append(f"第 {i} 行 Thickness 格式錯誤：{e}")
            try:
                if v_raw and v_raw not in ("N/A", "--"):
                    v_val = float(v_raw)
                    ttv.append(v_val)
            except ValueError:
                pass
            try:
                if p_raw and p_raw not in ("N/A", "--"):
                    p_val = int(float(p_raw))
            except ValueError:
                pass
            raw_rows.append({"wafer_id": wid, "thickness": t_val, "ttv": v_val, "particle": p_val})

        if not thickness and not ttv:
            return json_error("CSV 中找不到有效的 Thickness 或 TTV 資料，請確認欄位名稱", 400)
        particle_lots = []

    # ── JSON 模式 ─────────────────────────────────────────────
    else:
        body = json_body()
        batch_id = body.get("batch_id", "未命名批次")
        thickness = [float(v) for v in body.get("thickness", [])]
        ttv       = [float(v) for v in body.get("ttv", [])]
        spec      = body.get("spec", {})
        errors    = []
        particle_lots = body.get("particle_lots", [])

    spec.setdefault("thickness_usl", 705.0)
    spec.setdefault("thickness_lsl", 695.0)
    spec.setdefault("ttv_usl", 2.0)
    spec.setdefault("ttv_lsl", 0.0)

    try:
        result = run_all_charts(thickness, ttv, particle_lots, spec)
    except Exception as e:
        return json_error(f"SPC 計算失敗：{e}", 500)

    # 存入本地歷史（簡易 JSON 檔）
    history_path = BASE_DIR / "spc_history.json"
    history = read_json(history_path, [])
    import datetime
    entry = {
        "batch_id": batch_id,
        "analyzed_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "spec": spec,
        "thickness_n": len(thickness),
        "ttv_n": len(ttv),
        "thickness_cpk": result["summary"].get("thickness_cpk"),
        "ttv_cpk": result["summary"].get("ttv_cpk"),
        "needs_attention": result["summary"].get("needs_attention", False),
        "result": result,
        "parse_errors": errors,
        "raw_rows": raw_rows,   # 原始量測資料，供 FOSB 報表使用
    }
    history.insert(0, entry)
    write_json(history_path, history[:50])  # 保留最近 50 筆

    return jsonify({"success": True, "batch_id": batch_id, "result": result,
                    "parse_errors": errors})


@app.route('/api/spc/history', methods=['GET'])
def api_spc_history():
    """回傳最近 50 筆 SPC 分析記錄（不含完整 result，僅摘要）"""
    history_path = BASE_DIR / "spc_history.json"
    history = read_json(history_path, [])
    summary = [{
        "batch_id":       h.get("batch_id"),
        "analyzed_at":    h.get("analyzed_at"),
        "thickness_cpk":  h.get("thickness_cpk"),
        "ttv_cpk":        h.get("ttv_cpk"),
        "thickness_n":    h.get("thickness_n"),
        "ttv_n":          h.get("ttv_n"),
        "needs_attention": h.get("needs_attention", False),
        "parse_errors":   h.get("parse_errors", []),
        "has_raw_rows":   len(h.get("raw_rows", [])) > 0,
    } for h in history]
    return jsonify({"items": summary})


@app.route('/api/spc/fosb', methods=['GET'])
def api_spc_fosb():
    """
    產生 FOSB 客戶報表 Excel。
    Query param: batch_id
    FOSB 格式：最多 6 個 FOSB，每個約 74 wafer，
    欄位固定：wafer_id(C,I,O,U,AA,AG)，thickness(E,K,Q,W,AC,AI)，ttv(F,L,R,X,AD,AJ)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

    batch_id = request.args.get("batch_id", "")
    history_path = BASE_DIR / "spc_history.json"
    history = read_json(history_path, [])
    entry = next((h for h in history if h.get("batch_id") == batch_id), None)
    if not entry:
        return json_error(f"找不到批次 '{batch_id}'，請先上傳分析", 404)
    raw_rows = entry.get("raw_rows", [])
    if not raw_rows:
        return json_error("此批次無原始量測資料，請重新上傳 CSV 後再產生報表", 400)

    spec = entry.get("spec", {})
    t_usl = spec.get("thickness_usl", 705.0)
    t_lsl = spec.get("thickness_lsl", 695.0)
    v_usl = spec.get("ttv_usl", 2.0)

    # FOSB 欄位對應
    FOSB_MAP = [
        {"wafer_id": "C",  "thickness": "E",  "ttv": "F"},
        {"wafer_id": "I",  "thickness": "K",  "ttv": "L"},
        {"wafer_id": "O",  "thickness": "Q",  "ttv": "R"},
        {"wafer_id": "U",  "thickness": "W",  "ttv": "X"},
        {"wafer_id": "AA", "thickness": "AC", "ttv": "AD"},
        {"wafer_id": "AG", "thickness": "AI", "ttv": "AJ"},
    ]
    DATA_START_ROW = 5
    WAFERS_PER_FOSB = 74

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FOSB量測報表"

    # ── 樣式定義 ──
    hdr_font  = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    rej_fill  = PatternFill("solid", fgColor="FFCCCC")
    ok_fill   = PatternFill("solid", fgColor="E6F4EA")
    thin_side = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center    = Alignment(horizontal="center", vertical="center")

    # ── 第 1 行：報表標題 ──
    ws.merge_cells("A1:AJ1")
    ws["A1"] = f"潔沛企業有限公司 — FOSB 晶圓量測報表  批次：{batch_id}  產出：{entry.get('analyzed_at','')}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="0A1628")
    ws["A1"].font = Font(bold=True, size=12, color="7DD3FC")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 22

    # ── 第 2 行：SPC 摘要（獨立行，不與 FOSB 標頭衝突）──
    ws["A2"] = "Thickness Cpk:"
    ws["B2"] = entry.get("thickness_cpk") or "–"
    ws["D2"] = "TTV Cpk:"
    ws["E2"] = entry.get("ttv_cpk") or "–"
    ws["G2"] = f"Thickness USL/LSL: {t_usl}/{t_lsl} μm"
    ws["L2"] = f"TTV USL: {v_usl} μm"
    for cell_ref in ["A2","D2","G2","L2"]:
        ws[cell_ref].font = Font(bold=True, size=9, color="94A3B8")
    for cell_ref in ["B2","E2"]:
        v = ws[cell_ref].value
        try:
            color = "22C55E" if float(v) >= 1.33 else ("F59E0B" if float(v) >= 1.00 else "EF4444")
        except (TypeError, ValueError):
            color = "64748B"
        ws[cell_ref].font = Font(bold=True, size=11, color=color)
    ws.row_dimensions[2].height = 18

    # ── 第 3 行：FOSB 群組標頭 ──
    for fosb_i, cols in enumerate(FOSB_MAP):
        header_col_idx = column_index_from_string(cols["wafer_id"])
        ttv_col_idx = column_index_from_string(cols["ttv"])
        ws.merge_cells(start_row=3, start_column=header_col_idx, end_row=3, end_column=ttv_col_idx)
        c = ws.cell(row=3, column=header_col_idx)
        c.value = f"FOSB {fosb_i + 1}"
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
    ws.row_dimensions[3].height = 18

    # ── 第 4 行：欄位子標頭 ──
    for fosb_i, cols in enumerate(FOSB_MAP):
        for sub_label, col_letter in [("Wafer ID", cols["wafer_id"]),
                                       ("Thickness (μm)", cols["thickness"]),
                                       ("TTV (μm)", cols["ttv"])]:
            c = column_index_from_string(col_letter)
            cell = ws.cell(row=4, column=c)
            cell.value = sub_label
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2D5A8E")
            cell.alignment = center
            cell.border = thin_border
    ws.row_dimensions[4].height = 18

    # ── 資料列 ──
    for fosb_i, cols in enumerate(FOSB_MAP):
        start = fosb_i * WAFERS_PER_FOSB
        chunk = raw_rows[start: start + WAFERS_PER_FOSB]
        for row_i, row in enumerate(chunk):
            excel_row = DATA_START_ROW + row_i
            t_val = row.get("thickness")
            v_val = row.get("ttv")

            # Wafer ID
            wid_col = column_index_from_string(cols["wafer_id"])
            ws.cell(row=excel_row, column=wid_col).value = row.get("wafer_id", "")
            ws.cell(row=excel_row, column=wid_col).font = Font(size=9)
            ws.cell(row=excel_row, column=wid_col).border = thin_border

            # Thickness
            t_col = column_index_from_string(cols["thickness"])
            t_cell = ws.cell(row=excel_row, column=t_col)
            t_cell.value = round(t_val, 4) if t_val is not None else ""
            t_cell.number_format = "0.000"
            t_cell.font = Font(size=9)
            t_cell.border = thin_border
            if t_val is not None:
                t_oos = not (t_lsl <= t_val <= t_usl)
                t_cell.fill = rej_fill if t_oos else ok_fill

            # TTV
            v_col = column_index_from_string(cols["ttv"])
            v_cell = ws.cell(row=excel_row, column=v_col)
            v_cell.value = round(v_val, 4) if v_val is not None else ""
            v_cell.number_format = "0.000"
            v_cell.font = Font(size=9)
            v_cell.border = thin_border
            if v_val is not None:
                v_oos = not (0 <= v_val <= v_usl)
                v_cell.fill = rej_fill if v_oos else ok_fill

    # ── 欄寬 ──
    for cols in FOSB_MAP:
        ws.column_dimensions[cols["wafer_id"]].width = 15
        ws.column_dimensions[cols["thickness"]].width = 13
        ws.column_dimensions[cols["ttv"]].width = 10

    # ── 凍結標頭 ──
    ws.freeze_panes = f"A{DATA_START_ROW}"

    buf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(buf.name)
    buf.close()

    safe_batch = "".join(c if c.isalnum() or c in "-_." else "_" for c in batch_id)
    filename = f"FOSB_{safe_batch}.xlsx"
    return send_file(buf.name, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── JWT 身份驗證 API ─────────────────────────────────────────────────────────

USERS_PATH = BASE_DIR / "users.json"
JWT_EXPIRY_HOURS = 12

def _jwt_secret() -> str:
    return get_or_create_flask_secret()

def _load_users() -> list[dict]:
    return read_json(USERS_PATH, [])

def _hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def _make_token(user: dict) -> str:
    payload = {
        "sub": user["username"],
        "role": user["role"],
        "display": user["display"],
        "exp": time.time() + JWT_EXPIRY_HOURS * 3600,
    }
    return pyjwt.encode(payload, _jwt_secret(), algorithm="HS256")

def _decode_token(token: str) -> dict | None:
    try:
        return pyjwt.decode(token, _jwt_secret(), algorithms=["HS256"])
    except pyjwt.ExpiredSignatureError:
        return None
    except pyjwt.InvalidTokenError:
        return None

def require_auth(f):
    """裝飾器：驗證 JWT，失敗回傳 401。將解碼後的 payload 以 g.user 傳入。"""
    @wraps(f)
    def decorated(*args, **kwargs):
        from flask import g
        auth_header = request.headers.get("Authorization", "")
        token = auth_header.removeprefix("Bearer ").strip()
        if not token:
            return jsonify({"error": "未登入，請先登入系統"}), 401
        payload = _decode_token(token)
        if payload is None:
            return jsonify({"error": "登入已過期，請重新登入"}), 401
        g.user = payload
        return f(*args, **kwargs)
    return decorated

@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    body = json_body()
    username = (body.get("username") or "").strip().lower()
    password = body.get("password") or ""
    users = _load_users()
    user = next((u for u in users if u["username"] == username), None)
    if not user or user["password_hash"] != _hash_password(password):
        return jsonify({"success": False, "error": "帳號或密碼錯誤"}), 401
    token = _make_token(user)
    return jsonify({
        "success": True,
        "token": token,
        "user": {"username": user["username"], "role": user["role"], "display": user["display"]},
        "expires_in": JWT_EXPIRY_HOURS * 3600,
    })

@app.route('/api/auth/me', methods=['GET'])
@require_auth
def api_auth_me():
    from flask import g
    return jsonify({"success": True, "user": {
        "username": g.user["sub"],
        "role": g.user["role"],
        "display": g.user["display"],
    }})

@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def api_auth_change_password():
    from flask import g
    body = json_body()
    old_pw = body.get("old_password") or ""
    new_pw = body.get("new_password") or ""
    if len(new_pw) < 6:
        return json_error("新密碼至少需要 6 個字元")
    users = _load_users()
    user = next((u for u in users if u["username"] == g.user["sub"]), None)
    if not user or user["password_hash"] != _hash_password(old_pw):
        return json_error("舊密碼錯誤", 401)
    user["password_hash"] = _hash_password(new_pw)
    write_json(USERS_PATH, users)
    return jsonify({"success": True, "message": "密碼已更新"})


def kill_port(port: int) -> None:
    """啟動前清除佔用指定 port 的舊 Python 程序（Windows）"""
    import subprocess
    current_pid = os.getpid()
    try:
        result = subprocess.run(
            ['netstat', '-ano'], capture_output=True, text=True, timeout=5
        )
        for line in result.stdout.splitlines():
            if f':{port}' in line and 'LISTENING' in line:
                parts = line.split()
                if not parts:
                    continue
                try:
                    pid = int(parts[-1])
                except ValueError:
                    continue
                if pid != current_pid:
                    subprocess.run(
                        ['taskkill', '/F', '/PID', str(pid)],
                        capture_output=True, timeout=5
                    )
                    print(f'[cleanup] 已關閉舊程序 PID {pid}（port {port}）')
    except Exception:
        pass


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8888))
    host = os.environ.get('HOST', '127.0.0.1')
    kill_port(port)
    print(f'[server] http://{host}:{port}')
    print(f'[base] {BASE_DIR}')
    for warning in startup_security_warnings(host):
        print(f'[security-warning] {warning}')
    app.run(host=host, port=port, debug=False, use_reloader=False)
