# -*- coding: utf-8 -*-
"""
update_111_with_actuals.py
從 11.5 生產日報_目檢合一表單.xlsx 彙整每日實際投入/入庫/利用率，
回填到 11.1 生產計劃表.xlsx 各 WEEK 工作表，
輸出 11.1生產計劃表_更新版.xlsx（原格式保留，僅更新數值）。
"""
import os, re
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font

BASE   = r"C:\Users\USER\Desktop\NAS\公用\ISO文件建立\潔沛修訂版\11 生產作業管理程序\記錄"
F_115  = os.path.join(BASE, "11.5生產日報_目檢合一表單.xlsx")
F_111  = os.path.join(BASE, "11.1生產計劃表.xlsx")
OUTPUT = os.path.join(BASE, "11.1生產計劃表_更新版.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 工具函數
# ─────────────────────────────────────────────────────────────────────────────
def parse_num(v):
    if v is None: return 0
    if isinstance(v, float): return int(v) if v == int(v) else v
    if isinstance(v, int):   return v
    m = re.match(r'^(\d+\.?\d*)', str(v).strip())
    return int(float(m.group(1))) if m else 0


def normalise_date(v):
    """
    統一轉為 'YYYY/MM/DD' 字串，接受：
    - datetime 物件
    - '2026-01-05 00:00:00'
    - '2026/1/23(五)'  → '2026/01/23'
    - '2026/01/05'
    """
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y/%m/%d')
    s = str(v).strip()
    # '2026-01-05' or '2026-01-05 00:00:00'
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
    # '2026/1/23(五)' or '2026/01/23'
    m = re.match(r'(\d{4})/(\d{1,2})/(\d{1,2})', s)
    if m:
        return f"{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
    return None


def sheet_to_date(sheet_name):
    """'生產日報2026.01.05' → '2026/01/05'
    對多批次工作表（如 '2026.02.201' = 2月2日第1批）正確解析：
    用 (?!\\d) 確保日期部份恰好是 2 位數，後面沒有額外的數字。
    """
    m = re.search(r'(\d{3,4})\.(\d{2})\.(\d{2})(?!\d)', sheet_name)
    if not m:
        return None
    y = int(m.group(1))
    if y < 1911:
        y += 1911
    return f"{y}/{m.group(2)}/{m.group(3)}"


def any_to_date(v):
    """將任意日期格式轉為 'YYYY/MM/DD'，支援：
    - 民國年：'115.01.05' → '2026/01/05'
    - 西元點分：'2026.01.21' → '2026/01/21'
    - datetime 物件
    """
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y/%m/%d')
    s = str(v).strip()
    # 民國年 (2-3 digit)
    m = re.match(r'^(\d{2,3})\.(\d{2})\.(\d{2})$', s)
    if m:
        y = int(m.group(1)) + 1911
        return f"{y}/{m.group(2)}/{m.group(3)}"
    # 西元年點分 (4-digit year)
    m = re.match(r'^(\d{4})\.(\d{2})\.(\d{2})$', s)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return None

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1：彙整 11.5 每日實績
# by_date[date] = {
#   'ocr_input':  sum of OCR 投入數（第一道工序）
#   'aoi_good':   sum of AOI 良品數
#   'aoi_bad':    sum of AOI 不良品數
#   'vis_good':   sum of 目檢區 良品數（AOI 不足時備用）
#   'all_fosbs':  {fosb: input_qty}（去重計算實際投入備用）
# }
# ─────────────────────────────────────────────────────────────────────────────
print("STEP1: aggregating 11.5 by date ...")
by_date = defaultdict(lambda: {
    'ocr_input': 0, 'aoi_good': 0, 'aoi_bad': 0,
    'vis_good': 0,  'vis_bad': 0,
    'all_fosbs': {},
})

wb115 = openpyxl.load_workbook(F_115, data_only=True)
for sname in wb115.sheetnames:
    if '空白' in sname or '範本' in sname:
        continue
    ws115 = wb115[sname]
    all_rows = list(ws115.iter_rows(values_only=True))
    # 從工作表名稱嘗試取得日期（多批次工作表如 2026.02.201 可能回傳 None）
    sheet_d = sheet_to_date(sname)

    current_date    = sheet_d  # 可能為 None，後續從資料列補充
    current_station = ''

    for row in all_rows[4:]:
        if len(row) < 7:
            continue
        # 更新站點（合併儲存格延用）
        if row[1] and str(row[1]).strip():
            current_station = str(row[1]).strip()
        # 更新日期（支援民國年 115.xx.xx 與西元年 2026.xx.xx）
        if row[0]:
            d = any_to_date(row[0])
            if d:
                current_date = d
        # 日期未知的列跳過
        if current_date is None:
            continue

        fosb      = row[3]
        input_qty = row[4]
        good_qty  = row[6]
        bad_qty   = row[7]

        if fosb is None and input_qty is None:
            continue

        input_n = parse_num(input_qty)
        good_n  = parse_num(good_qty)
        bad_n   = parse_num(bad_qty)

        d = by_date[current_date]

        if current_station == 'OCR':
            d['ocr_input'] += input_n
        elif current_station == 'AOI':
            d['aoi_good'] += good_n
            d['aoi_bad']  += bad_n
        elif current_station == '目檢區':
            d['vis_good'] += good_n
            d['vis_bad']  += bad_n

        # 所有批號 → 去重計算實際投入（備用）
        fosb_str = str(fosb).strip() if fosb else f'_row{id(row)}'
        if fosb_str not in d['all_fosbs']:
            d['all_fosbs'][fosb_str] = input_n

wb115.close()
print(f"  -> {len(by_date)} distinct dates found in 11.5")

def get_actual(date_str):
    """實際投入 = OCR 小計；若無則用去重批號合計"""
    d = by_date.get(date_str, {})
    v = d.get('ocr_input', 0)
    if v == 0:
        v = sum(d.get('all_fosbs', {}).values())
    return v or None

def get_inku(date_str):
    """入庫良品 = AOI 良品；若無則用目檢區"""
    d = by_date.get(date_str, {})
    v = d.get('aoi_good', 0)
    if v == 0:
        v = d.get('vis_good', 0)
    return v or None

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2：開啟 11.1，更新各 WEEK 工作表
# 固定列號（1-based openpyxl）：
#   row 3 = 日期表頭
#   row 4 = 計劃投入
#   row 5 = 實際投入   ← 更新
#   row 6 = 入庫數量   ← 更新
#   row 7 = 產能利用率  ← 更新
#   資料欄起始 = 第 10 欄 (index 9, 1-based col=10)
# ─────────────────────────────────────────────────────────────────────────────
print("STEP2: updating 11.1 ...")
wb111 = openpyxl.load_workbook(F_111)   # no data_only so we can write

# 良率標色樣式
RED_FILL = PatternFill("solid", fgColor="FF9999")
YLW_FILL = PatternFill("solid", fgColor="FFFACD")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")

updated_sheets = []

for sname in wb111.sheetnames:
    # 只處理 WEEK 工作表（有 2026 日期的週報）
    if 'WEEK' not in sname.upper() and 'week' not in sname.lower():
        continue

    ws = wb111[sname]
    print(f"  Processing [{sname}] ...")

    # ── 找出所有日期欄位（row 3, 從 col 10 開始）────────────────────────────
    date_cols = {}   # {col_1based: date_str}
    for col in range(10, ws.max_column + 1):
        v = ws.cell(row=3, column=col).value
        if v is None:
            break
        date_str = normalise_date(v)
        if date_str:
            date_cols[col] = date_str

    if not date_cols:
        print(f"    WARN: no date columns found, skip")
        continue

    # ── 逐欄回填────────────────────────────────────────────────────────────
    total_plan   = 0
    total_actual = 0
    total_inku   = 0
    date_count   = 0

    for col, date_str in date_cols.items():
        plan_cell   = ws.cell(row=4, column=col)
        actual_cell = ws.cell(row=5, column=col)
        inku_cell   = ws.cell(row=6, column=col)
        util_cell   = ws.cell(row=7, column=col)

        plan_qty   = parse_num(plan_cell.value)
        actual_qty = get_actual(date_str)
        inku_qty   = get_inku(date_str)

        if actual_qty is None and inku_qty is None:
            # 此日期在 11.5 中無資料 → 保留原值
            total_plan   += plan_qty
            total_actual += parse_num(actual_cell.value)
            total_inku   += parse_num(inku_cell.value)
            date_count   += 1
            continue

        actual_qty = actual_qty or 0
        inku_qty   = inku_qty   or 0

        # 計算利用率（相對計劃投入）
        if plan_qty > 0:
            util_rate = inku_qty / plan_qty
        elif actual_qty > 0:
            util_rate = inku_qty / actual_qty
        else:
            util_rate = None

        # 寫入數值
        actual_cell.value = actual_qty
        inku_cell.value   = inku_qty
        if util_rate is not None:
            util_cell.value = util_rate
            util_cell.number_format = '0.0%'
            # 良率色碼
            if util_rate < 0.90:
                util_cell.fill = RED_FILL
            elif util_rate < 0.95:
                util_cell.fill = YLW_FILL
            else:
                util_cell.fill = GRN_FILL

        total_plan   += plan_qty
        total_actual += actual_qty
        total_inku   += inku_qty
        date_count   += 1

        print(f"    {date_str}: plan={plan_qty} actual={actual_qty} inku={inku_qty}"
              + (f" util={util_rate:.1%}" if util_rate is not None else " util=—"))

    # ── 更新合計欄（若存在）──────────────────────────────────────────────
    # 合計欄位置 = 最後日期欄的下一欄
    last_date_col = max(date_cols.keys())
    sum_col = last_date_col + 1
    # 確認合計欄：Row4 有值，且值約等於計劃合計
    sum_check = ws.cell(row=4, column=sum_col).value
    has_sum_col = (sum_check is not None and
                   abs(parse_num(sum_check) - total_plan) < 10)

    if has_sum_col:
        actual_sum_cell = ws.cell(row=5, column=sum_col)
        inku_sum_cell   = ws.cell(row=6, column=sum_col)
        util_sum_cell   = ws.cell(row=7, column=sum_col)
        actual_sum_cell.value = total_actual
        inku_sum_cell.value   = total_inku
        if total_plan > 0:
            u = total_inku / total_plan
            util_sum_cell.value = u
            util_sum_cell.number_format = '0.0%'
            util_sum_cell.fill = RED_FILL if u < 0.90 else (YLW_FILL if u < 0.95 else GRN_FILL)
        print(f"  Total: plan={total_plan} actual={total_actual}"
              f" inku={total_inku}"
              + (f" util={total_inku/total_plan:.1%}" if total_plan > 0 else ""))

    updated_sheets.append(sname)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3：儲存
# ─────────────────────────────────────────────────────────────────────────────
wb111.save(OUTPUT)
print(f"\nOK: Updated {len(updated_sheets)} sheets: {updated_sheets}")
print(f"OK: Saved -> {OUTPUT}")
print("DONE")
