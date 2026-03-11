"""
generate_record.py — 依據資料生成已填寫的 XLSX 記錄表格

支援類型：
  env        → 6 工作環境管理程序 / 表單 / 6.1環境監控記錄表.xlsx
  production → 11 生產作業管理程序 / 表單 / 11.5生產日報表.xlsx
  quality    → 12 品質檢驗管理程序 / 表單 / 12.1品質管理記錄表(原料).xlsx

呼叫方式（由 server.py 引用）：
  out_path = generate_record.run(rec_type, data_list, out_path)

直接執行（測試）：
  python generate_record.py env
"""

import shutil
import json
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    openpyxl = None

BASE = Path(__file__).parent.resolve()

# ── 範本路徑 ─────────────────────────────────────────────────
TEMPLATES = {
    'env':        BASE / '6 工作環境管理程序'     / '表單' / '6.1環境監控記錄表.xlsx',
    'production': BASE / '11 生產作業管理程序'    / '表單' / '11.5生產日報表.xlsx',
    'quality':    BASE / '12 品質檢驗管理程序'    / '表單' / '12.1品質管理記錄表(原料).xlsx',
}
# 如果表單不存在，從記錄目錄取最新範本
FALLBACK_TEMPLATES = {
    'env':        BASE / '6 工作環境管理程序'     / '記錄' / '6.1環境監控記錄表.xlsx',
    'production': BASE / '11 生產作業管理程序'    / '記錄' / '11.5生產日報_目檢合一表單.xlsx',
    'quality':    BASE / '12 品質檢驗管理程序'    / '記錄' / '12.1潔沛品質管理記錄表(原料).xlsx',
}

# 邊框樣式
_thin = Side(style='thin', color='000000')
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _apply_cell(ws, row, col, value, bold=False, center=False, fill_hex=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if fill_hex:
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_hex)
    cell.border = _border
    return cell


# ════════════════════════════════════════════════════════════
# env — 環境監控記錄表
# ════════════════════════════════════════════════════════════

def _fill_env(ws, data: list):
    """
    data 結構（從儀表板 envRecords 傳入）:
      [{ date, temp, humidity, pressure, result, operator, note }, ...]
    範本欄位（Row 3）:
      A=日期, B=實際溫度, C=溫度合格, D=濕度, E=濕度合格, F=壓差, G=壓差合格, H=備註
    """
    START_ROW = 4   # 資料從第 4 行開始
    HEADER_ROW = 3

    # 確認表頭
    headers = ['日期', '實際溫度(°C)(15~25)', '溫度是否合格',
               '濕度(%RH)(40~70)', '濕度是否合格',
               '壓差(Pa)(0.6~1.5)', '壓差是否合格', '備註']
    for col, h in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col, value=h).font = Font(bold=True)

    # 合規判定
    def ok(val, lo, hi):
        if val is None or val == '':
            return ''
        try:
            v = float(str(val).split('/')[0])
            return 'Y' if lo <= v <= hi else 'N'
        except Exception:
            return ''

    for i, row in enumerate(data):
        r = START_ROW + i
        date_str = row.get('date', '')
        temp = row.get('temp', '')
        humidity = row.get('humidity', '')
        pressure = row.get('pressure', '')
        note = row.get('note', row.get('operator', ''))

        temp_ok = ok(temp, 15, 25)
        hum_ok  = ok(humidity, 40, 70)
        prs_ok  = ok(pressure, 0.6, 1.5)

        vals = [date_str, temp, temp_ok, humidity, hum_ok, pressure, prs_ok, note]
        fill = 'FFF2CC' if 'N' in (temp_ok, hum_ok, prs_ok) else None
        for col, val in enumerate(vals, 1):
            _apply_cell(ws, r, col, val, fill_hex=fill)

    return ws


# ════════════════════════════════════════════════════════════
# production — 生產日報記錄表
# ════════════════════════════════════════════════════════════

def _fill_production(ws, data: list):
    """
    data 結構:
      [{ lot, customer, product, input, good, defect, yieldRate,
         defectReasons, operator, note }, ...]
    範本欄位（Row 3）:
      A=批次號碼, B=客戶代號/產品名稱, C=投入數, D=良品數, E=不良品數,
      F=生產人員/確認, G=備註
    """
    START_ROW = 4
    HEADER_ROW = 3

    headers = ['批次號碼', '客戶代號/產品名稱', '投入數', '良品數',
               '不良品數', '良率(%)', '生產人員', '備註']
    for col, h in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col, value=h).font = Font(bold=True)

    total_input = 0; total_good = 0
    for i, row in enumerate(data):
        r = START_ROW + i
        lot = row.get('lot', row.get('waferBoatLot', ''))
        customer = row.get('customer', row.get('clientCode', ''))
        product = row.get('product', row.get('productName', ''))
        cust_prod = f"{customer} / {product}" if product else customer
        inp  = row.get('input', row.get('inputQty', 0)) or 0
        good = row.get('good', row.get('goodQty', 0)) or 0
        defect = row.get('defect', row.get('defectQty', 0)) or 0
        yr = row.get('yieldRate', round(good / inp * 100, 1) if inp else '')
        operator = row.get('operator', '')
        reasons = row.get('defectReasons', [])
        if isinstance(reasons, list):
            reasons_str = ', '.join(str(x) for x in reasons if x)
        else:
            reasons_str = str(reasons)
        note = row.get('note', reasons_str)

        total_input += int(inp) if inp else 0
        total_good  += int(good) if good else 0

        low_yield = isinstance(yr, (int, float)) and yr < 90
        fill = 'FFE0E0' if low_yield else None
        vals = [lot, cust_prod, inp, good, defect, yr, operator, note]
        for col, val in enumerate(vals, 1):
            _apply_cell(ws, r, col, val, fill_hex=fill)

    # 合計行
    total_row = START_ROW + len(data)
    total_yr = round(total_good / total_input * 100, 1) if total_input else ''
    _apply_cell(ws, total_row, 1, '合計', bold=True, fill_hex='D9D9D9')
    _apply_cell(ws, total_row, 3, total_input, bold=True, fill_hex='D9D9D9')
    _apply_cell(ws, total_row, 4, total_good, bold=True, fill_hex='D9D9D9')
    _apply_cell(ws, total_row, 5, total_input - total_good, bold=True, fill_hex='D9D9D9')
    _apply_cell(ws, total_row, 6, total_yr, bold=True, fill_hex='D9D9D9')

    return ws


# ════════════════════════════════════════════════════════════
# quality — 品質管理記錄表（原料）
# ════════════════════════════════════════════════════════════

def _fill_quality(ws, data: list):
    """
    data 結構:
      [{ materialName, batchNo, quantity, spec, inspQty,
         ph, density, ri, rotation, result, note }, ...]
    範本欄位（Row 5）:
      A=原料名稱, B=原料批號, C=原料數量, D=規格, E=品檢數量,
      F=PH值, G=比重值, H=RI值, I=旋光度, J=結果, K=備註
    """
    START_ROW = 6
    HEADER_ROW = 5

    headers = ['原料名稱', '原料批號', '原料數量', '規格', '品檢數量',
               'PH值', '比重值', 'RI值', '旋光度', '檢驗結果', '備註']
    for col, h in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col, value=h).font = Font(bold=True)

    for i, row in enumerate(data):
        r = START_ROW + i
        result = row.get('result', row.get('inspResult', ''))
        fill = 'FFE0E0' if result in ('不合格', 'NG', 'FAIL') else None
        vals = [
            row.get('materialName', row.get('name', '')),
            row.get('batchNo', row.get('batch', '')),
            row.get('quantity', row.get('qty', '')),
            row.get('spec', ''),
            row.get('inspQty', ''),
            row.get('ph', ''),
            row.get('density', ''),
            row.get('ri', ''),
            row.get('rotation', ''),
            result,
            row.get('note', ''),
        ]
        for col, val in enumerate(vals, 1):
            _apply_cell(ws, r, col, val, fill_hex=fill)

    return ws


# ════════════════════════════════════════════════════════════
# 主函式
# ════════════════════════════════════════════════════════════

def run(rec_type: str, data: list, out_path: str) -> str:
    """
    rec_type: 'env' | 'production' | 'quality'
    data:     list of dicts（從前端 JSON 傳入）
    out_path: 輸出 XLSX 路徑
    回傳: out_path（供 send_file 使用）
    """
    if openpyxl is None:
        raise ImportError('請先安裝 openpyxl: pip install openpyxl')

    # 選擇範本
    tmpl = TEMPLATES.get(rec_type)
    if not tmpl or not tmpl.exists():
        tmpl = FALLBACK_TEMPLATES.get(rec_type)
    if not tmpl or not tmpl.exists():
        # 全新空白工作簿
        tmpl = None

    if tmpl:
        shutil.copy2(tmpl, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        # 清除舊資料（從第 4 行以後），保留標題行；跳過合併儲存格
        clear_from = 4 if rec_type in ('env', 'production') else 6
        max_row = ws.max_row
        for row_idx in range(clear_from, max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None
                except AttributeError:
                    pass  # MergedCell — skip
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = {'env': '環境監控', 'production': '生產日報', 'quality': '品質管理'}.get(rec_type, '記錄')
        # 基本標題
        ws.cell(row=1, column=1, value='潔沛企業有限公司').font = Font(bold=True, size=14)
        titles = {'env': '環境監控記錄表', 'production': '生產日報記錄表', 'quality': '品質管理記錄表（原料）'}
        ws.cell(row=2, column=1, value=titles.get(rec_type, '記錄')).font = Font(bold=True, size=12)
        # 生成時間
        ws.cell(row=1, column=8, value=f"生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # 填入資料
    fill_funcs = {
        'env':        _fill_env,
        'production': _fill_production,
        'quality':    _fill_quality,
    }
    fill_func = fill_funcs.get(rec_type)
    if fill_func:
        fill_func(ws, data)

    # 自動調整欄寬（跳過合併儲存格）
    from openpyxl.utils import get_column_letter
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and not hasattr(cell, 'column') is False:
                try:
                    col_idx = cell.column
                    clen = len(str(cell.value))
                    if clen > col_widths.get(col_idx, 0):
                        col_widths[col_idx] = clen
                except Exception:
                    pass
    for col_idx, max_len in col_widths.items():
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 40)
        except Exception:
            pass

    wb.save(out_path)
    return out_path


# ════════════════════════════════════════════════════════════
# 直接執行測試
# ════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys, os

    rec_type = sys.argv[1] if len(sys.argv) > 1 else 'env'
    test_data = {
        'env': [
            {'date': '2026-03-01', 'temp': '22.5', 'humidity': '55', 'pressure': '1.2', 'result': '合格'},
            {'date': '2026-03-02', 'temp': '28.0', 'humidity': '75', 'pressure': '0.4', 'result': '不合格'},
        ],
        'production': [
            {'lot': 'LOT001', 'customer': 'C001', 'product': '玻璃清洗', 'input': 100, 'good': 95, 'defect': 5},
            {'lot': 'LOT002', 'customer': 'C002', 'product': '薄片清洗', 'input': 200, 'good': 180, 'defect': 20},
        ],
        'quality': [
            {'materialName': '清洗劑A', 'batchNo': 'BL001', 'quantity': '20L', 'result': '合格'},
            {'materialName': '清洗劑B', 'batchNo': 'BL002', 'quantity': '10L', 'result': '不合格'},
        ],
    }

    out = BASE / f'_test_output_{rec_type}.xlsx'
    run(rec_type, test_data[rec_type], str(out))
    print(f'[OK] 輸出: {out}')
