from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "工作環境監控上傳模板.xlsx"


def build_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "環境監控資料"

    headers = [
        "Date",
        "DateTime",
        "Point",
        "Location",
        "0.3um",
        "0.5um",
        "5.0um",
        "Temp",
        "Humidity",
        "Pressure",
        "Operator",
        "Result",
        "Note",
    ]
    notes = [
        "日期，格式 YYYY-MM-DD",
        "量測時間，格式 YYYY-MM-DD HH:MM:SS",
        "點位，填 1~14",
        "地點，可留白；留白時系統會自動用點位產生",
        "0.3um 粒子數",
        "0.5um 粒子數",
        "5.0um 粒子數",
        "溫度，可留白",
        "濕度，可留白",
        "正壓，可留白",
        "記錄者",
        "可留白，系統會自動判定",
        "備註，可留白",
    ]
    sample_rows = [
        ["2025-11-14", "2025-11-14 11:22:50", "1", "", 16, 10, 3, "", "", "", "王小明", "", "粒子計數器範例"],
        ["2025-11-14", "2025-11-14 11:24:06", "2", "", 12, 7, 2, "", "", "", "王小明", "", ""],
        ["2025-11-14", "2025-11-14 13:30:00", "", "A區溫濕度測點", "", "", "", 22.3, 46.5, 12.0, "陳小華", "", "溫濕度/壓差範例"],
    ]

    fill_header = PatternFill("solid", fgColor="0F766E")
    fill_note = PatternFill("solid", fgColor="ECFEFF")
    fill_sample = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    ws.append(notes)
    for row in sample_rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for cell in ws[2]:
        cell.font = Font(color="334155", italic=True)
        cell.fill = fill_note
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill_sample
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"
    widths = {
        "A": 14,
        "B": 22,
        "C": 10,
        "D": 24,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 14,
        "L": 12,
        "M": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    guide = wb.create_sheet("填寫說明")
    guide["A1"] = "工作環境監控上傳模板說明"
    guide["A1"].font = Font(size=14, bold=True)
    guide["A3"] = "1. 請在「環境監控資料」工作表從第 3 列開始填資料。"
    guide["A4"] = "2. Date 請填日期；DateTime 請填完整日期時間，系統會依 DateTime 排序。"
    guide["A5"] = "3. Point 建議填 1~14；Location 可留白，系統會自動顯示為粒子計數點。"
    guide["A6"] = "4. 粒子欄位請填 0.3um / 0.5um / 5.0um。"
    guide["A7"] = "5. 溫度、濕度、正壓如沒有資料可留白。"
    guide["A8"] = "6. Result 可留白，系統會依規則自動判定。"
    guide["A10"] = "建議流程：手寫 -> 整理成這份 Excel -> 上傳到系統。"
    guide.column_dimensions["A"].width = 90

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    build_template(OUTPUT_PATH)
    print(OUTPUT_PATH)
