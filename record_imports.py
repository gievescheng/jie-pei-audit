from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent.resolve()

PRODUCTION_PREFIX = "11.5"
QUALITY_PREFIX = "12.1"

PRODUCTION_HEADER_ANCHOR = "Wafer Boat Lot"
PRODUCTION_SKIP_MARKERS = {
    "\u5be9\u6838\uff1a",  # 審核：
    "\u6838\u51c6\uff1a",  # 核准：
}
QUALITY_SKIP_MARKERS = {
    "\u54c1\u8cea\u6aa2\u9a57\u7d50\u679c",  # 品質檢驗結果
    "\u6838\u51c6\uff1a",  # 核准：
}


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _int_value(value):
    text = _text(value).replace(",", "")
    if not text:
        return 0
    try:
        return int(float(text))
    except Exception:
        return 0


def _yield_value(value):
    text = _text(value).replace("%", "").replace(",", "")
    if not text:
        return ""
    try:
        numeric = float(text)
        if 0 < numeric <= 1.2:
            numeric *= 100
        return round(numeric, 1)
    except Exception:
        return ""


def _split_customer_product(value: str) -> tuple[str, str]:
    text = _text(value)
    if "/" in text:
        customer, product = text.split("/", 1)
        return customer.strip(), product.strip()
    return "", text


def _scan_workbooks(prefix: str) -> list[Path]:
    return [
        path
        for path in BASE_DIR.rglob(f"{prefix}*.xlsx")
        if not path.name.startswith("~$")
    ]


def _choose_best(candidates: list[Path], parser) -> tuple[list[dict], str]:
    best_records: list[dict] = []
    best_path = ""
    best_score = (-1, -1)
    for path in candidates:
        records = parser(path)
        score = (len(records), len(path.name))
        if score > best_score:
            best_records = records
            best_path = str(path)
            best_score = score
    return best_records, best_path


def _find_production_header_row(ws) -> int | None:
    max_col = min(ws.max_column, 16)
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [_text(ws.cell(row_idx, col).value) for col in range(1, max_col + 1)]
        if PRODUCTION_HEADER_ANCHOR in values:
            return row_idx
    return None


def parse_production_record_file(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        header_row = _find_production_header_row(ws)
        if header_row is None:
            return []

        records: list[dict] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = [_text(ws.cell(row_idx, col).value) for col in range(1, 13)]
            if not any(row[:11]):
                continue
            if row[0] in PRODUCTION_SKIP_MARKERS or row[1] in PRODUCTION_SKIP_MARKERS:
                continue
            if not row[2] and not row[3]:
                continue
            if not any(ch.isdigit() for ch in row[0]) and not any(ch.isdigit() for ch in row[3]):
                continue

            customer, product = _split_customer_product(row[2])
            site = row[1]
            defect_reasons = [
                part.strip()
                for part in row[9].replace("\uff1f", ",").replace("?", ",").split(",")
                if part.strip()
            ]
            note_parts = []
            if site:
                note_parts.append("Site: " + site)
            if row[11]:
                note_parts.append(row[11])

            records.append(
                {
                    "date": row[0].replace(".", "-").replace("/", "-"),
                    "site": site,
                    "lot": row[3],
                    "customer": customer,
                    "product": product,
                    "input": _int_value(row[4]),
                    "good": _int_value(row[6]),
                    "defect": _int_value(row[7]),
                    "yieldRate": _yield_value(row[8]),
                    "defectReasons": defect_reasons,
                    "operator": row[10],
                    "note": " / ".join(part for part in note_parts if part),
                }
            )
        return records
    finally:
        wb.close()


def _is_quality_footer_row(row: list[str]) -> bool:
    leading = row[0]
    second = row[1]
    if leading in QUALITY_SKIP_MARKERS:
        return True
    if second in QUALITY_SKIP_MARKERS:
        return True
    if leading.startswith("\u6838\u51c6") or second.startswith("\u6838\u51c6"):
        return True
    if leading.startswith("\u54c1\u8cea\u6aa2\u9a57\u7d50\u679c"):
        return True
    return False


def parse_quality_record_file(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        header_row = 5
        if ws.max_row <= header_row:
            return []

        records: list[dict] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = [_text(ws.cell(row_idx, col).value) for col in range(1, 12)]
            if not any(row[:10]):
                continue
            if _is_quality_footer_row(row):
                continue
            if not row[0] and not row[1]:
                continue

            appearance = row[9].upper()
            result = "PASS" if appearance in {"OK", "PASS"} else appearance or "PASS"
            records.append(
                {
                    "materialName": row[0],
                    "batchNo": row[1],
                    "quantity": row[2],
                    "spec": row[3],
                    "inspQty": row[4],
                    "ph": row[5],
                    "density": row[6],
                    "ri": row[7],
                    "rotation": row[8],
                    "result": result,
                    "note": row[10],
                }
            )
        return records
    finally:
        wb.close()


def load_existing_production_records() -> tuple[list[dict], str]:
    return _choose_best(_scan_workbooks(PRODUCTION_PREFIX), parse_production_record_file)


def load_existing_quality_records() -> tuple[list[dict], str]:
    return _choose_best(_scan_workbooks(QUALITY_PREFIX), parse_quality_record_file)


def load_uploaded_production_records(path: Path) -> tuple[list[dict], str]:
    return parse_production_record_file(path), str(path)


def load_uploaded_quality_records(path: Path) -> tuple[list[dict], str]:
    return parse_quality_record_file(path), str(path)
