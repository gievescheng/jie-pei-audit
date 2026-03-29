# -*- coding: utf-8 -*-
"""
generate_11_1_supplement.py
從 11.5 生產日報_目檢合一表單.xlsx 抽取資料，
產生 11.1生產計劃表_補充版.xlsx（3 張工作表）
"""
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 路徑設定
# ─────────────────────────────────────────────────────────────────────────────
BASE   = r"C:\Users\USER\Desktop\NAS\公用\ISO文件建立\潔沛修訂版\11 生產作業管理程序\記錄"
F_115  = os.path.join(BASE, "11.5生產日報_目檢合一表單.xlsx")
F_111  = os.path.join(BASE, "11.1生產計劃表.xlsx")
OUTPUT = os.path.join(BASE, "11.1生產計劃表_補充版.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 工具函數
# ─────────────────────────────────────────────────────────────────────────────
def parse_num(v):
    """將 '25片' / 25 / 25.0 / None 統一轉成 int or None"""
    if v is None:
        return None
    if isinstance(v, float):
        return int(v) if v == int(v) else v
    if isinstance(v, int):
        return v
    m = re.match(r'^(\d+\.?\d*)', str(v).strip())
    return int(float(m.group(1))) if m else None


def sheet_to_date(sheet_name):
    """
    '生產日報2026.01.05' → '2026/01/05'
    支援民國年: '生產日報115.01.05' → '2026/01/05'
    """
    m = re.search(r'(\d{3,4})\.(\d{2})\.(\d{2})$', sheet_name)
    if not m:
        return sheet_name
    y = int(m.group(1))
    if y < 1911:          # 民國年
        y += 1911
    return f"{y}/{m.group(2)}/{m.group(3)}"


def roc_cell_to_date(v):
    """
    民國日期字串 '115.01.05' → '2026/01/05'
    如果 v 已是 datetime 物件就直接格式化
    """
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y/%m/%d')
    m = re.match(r'^(\d{2,3})\.(\d{2})\.(\d{2})$', str(v).strip())
    if m:
        y = int(m.group(1)) + 1911
        return f"{y}/{m.group(2)}/{m.group(3)}"
    return None

# ─────────────────────────────────────────────────────────────────────────────
# 樣式常數
# ─────────────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")   # 深藍
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")   # 淡藍
RED_FILL = PatternFill("solid", fgColor="FF9999")   # 紅（良率 < 90%）
YLW_FILL = PatternFill("solid", fgColor="FFFACD")   # 黃（良率 90–94%）
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")   # 綠（良率 >= 95%）
BLU_SEC  = PatternFill("solid", fgColor="2E75B6")   # 藍（段落標題）

_thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def hcell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = CENTER; c.border = BORDER
    return c


def dcell(ws, row, col, val, align=LEFT, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.alignment = align; c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1：讀取 11.5，逐張工作表解析
# ─────────────────────────────────────────────────────────────────────────────
print("STEP1: loading 11.5 ...")
wb115 = openpyxl.load_workbook(F_115, data_only=True)
records = []   # list of dict

for sname in wb115.sheetnames:
    if '空白' in sname or '範本' in sname:
        continue   # 跳過範本頁

    ws = wb115[sname]
    sheet_date = sheet_to_date(sname)

    all_rows = list(ws.iter_rows(values_only=True))

    # ── 取得訂單編號（row index 2, col index 9）────────────────────────────
    order_no = None
    if len(all_rows) > 2 and len(all_rows[2]) > 9:
        v = all_rows[2][9]
        if v is not None:
            order_no = str(v).strip()

    # ── 資料列從 index 4 開始 ────────────────────────────────────────────
    if len(all_rows) < 5:
        continue

    current_date    = sheet_date   # 如果 col0 有值就覆蓋
    current_station = ''           # 若站點欄是合併儲存格，向下延用

    for row in all_rows[4:]:
        if len(row) < 4:
            continue

        # col0: 日期（可能是民國年字串或 datetime，也可能為 None）
        if row[0] is not None:
            converted = roc_cell_to_date(row[0])
            if converted:
                current_date = converted

        # col1: 設施站點（合併儲存格時為 None，向下延用前一值）
        raw_station = row[1] if len(row) > 1 else None
        if raw_station is not None and str(raw_station).strip():
            current_station = str(raw_station).strip()

        station   = current_station
        product   = row[2]  if len(row) > 2  else None
        fosb      = row[3]  if len(row) > 3  else None
        input_qty = row[4]  if len(row) > 4  else None
        filtered  = row[5]  if len(row) > 5  else None
        good_qty  = row[6]  if len(row) > 6  else None
        bad_qty   = row[7]  if len(row) > 7  else None
        yield_v   = row[8]  if len(row) > 8  else None
        defect    = row[9]  if len(row) > 9  else None
        operator  = row[10] if len(row) > 10 else None
        remark    = row[11] if len(row) > 11 else None

        # 跳過空白列（站點和批號都空）
        if not station and fosb is None:
            continue
        # 跳過圖例列（僅有站點名稱，無投入量也無批號）
        if fosb is None and input_qty is None:
            continue

        # 解析數值
        input_n = parse_num(input_qty)
        good_n  = parse_num(good_qty)
        bad_n   = parse_num(bad_qty)

        # 計算良率：優先用好品/壞品計算（更準確），若無則讀欄位值
        if good_n is not None and bad_n is not None and (good_n + bad_n) > 0:
            yield_pct = good_n / (good_n + bad_n)
        elif isinstance(yield_v, (int, float)) and 0 < float(yield_v) <= 1:
            yield_pct = float(yield_v)
        elif isinstance(yield_v, str) and '%' in yield_v:
            try:
                yield_pct = float(yield_v.replace('%', '').strip()) / 100
            except ValueError:
                yield_pct = None
        elif good_n is not None and good_n > 0 and (bad_n is None or bad_n == 0):
            # 只有良品數，無不良品 → 100%
            yield_pct = 1.0
        else:
            yield_pct = None

        records.append({
            'date':      current_date,
            'order_no':  order_no or '',
            'product':   str(product).strip()  if product  else '',
            'station':   str(station).strip()  if station  else '',
            'fosb':      str(fosb).strip()     if fosb     else '',
            'input':     input_n,
            'filtered':  parse_num(filtered),
            'good':      good_n,
            'bad':       bad_n,
            'yield_pct': yield_pct,
            'defect':    str(defect).strip()   if defect   else '',
            'operator':  str(operator).strip() if operator else '',
            'remark':    str(remark).strip()   if remark   else '',
            'sheet':     sname,
        })

wb115.close()
print(f"  -> {len(records)} data rows parsed from {len(wb115.sheetnames)-1} sheets")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2：讀取 11.1，建立訂單→計劃量 查詢表
# ─────────────────────────────────────────────────────────────────────────────
print("STEP2: loading 11.1 ...")
order_plan = {}   # order_no_str -> {planned_qty, planned_date, process}
try:
    wb111 = openpyxl.load_workbook(F_111, data_only=True)
    for sname in wb111.sheetnames:
        ws = wb111[sname]
        rows111 = list(ws.iter_rows(values_only=True))

        # 11.1 結構（橫式，固定格式）：
        #   Row index 2: 表頭列 [null, 項次, 客戶代碼, 產品名稱, 訂單編號, 訂單數量, 預計交期, 產能, 日期, ...]
        #   Row index 3: [null, 1, C001, 產品名, null, "200片", null, "片/日", "計劃投入數量", ...]
        #   Row index 4: [null, null, null, null, 4515994888, null, "2025-11-14", null, "實際投入數量", ...]
        # → 訂單編號在 col 4（index 4）出現在 "實際投入數量" 那列
        # → 訂單數量（計劃量）在 col 5（index 5）出現在 "計劃投入數量" 那列

        planned_qty   = None
        order_no_val  = None
        deadline      = None
        product_name  = None

        def _try_order(v):
            """若 v 是大數字（訂單編號格式），回傳字串；否則 None"""
            if v is None:
                return None
            try:
                n = int(float(str(v)))
                if n >= 100000000:   # >= 9 位數
                    return str(n)
            except (ValueError, TypeError):
                pass
            return None

        def _try_date(v):
            if v is None:
                return ''
            if hasattr(v, 'strftime'):
                return v.strftime('%Y/%m/%d')
            s = str(v).strip()
            # "2026-01-12 00:00:00" → "2026/01/12"
            m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
            return f"{m.group(1)}/{m.group(2)}/{m.group(3)}" if m else s

        for row in rows111:
            # 抓取 col4 的訂單編號（只要是大數字就記）
            if len(row) > 4:
                ov = _try_order(row[4])
                if ov:
                    order_no_val = ov

            if len(row) < 9:
                continue
            label = str(row[8]).strip() if row[8] is not None else ''

            if '計劃投入' in label:
                # 訂單數量 = col 5
                if len(row) > 5 and row[5] is not None:
                    pq = parse_num(row[5])
                    if pq:
                        planned_qty = pq
                # 產品名稱 = col 3
                if len(row) > 3 and row[3]:
                    product_name = str(row[3]).strip()
                # WEEK 格式：訂單編號在此列 col 4，預計交期在 col 6
                if len(row) > 4:
                    ov = _try_order(row[4])
                    if ov:
                        order_no_val = ov
                if len(row) > 6 and row[6] and not deadline:
                    deadline = _try_date(row[6])

            elif '實際投入' in label:
                # 舊格式：訂單編號在此列 col 4，預計交期在 col 6
                if len(row) > 4:
                    ov = _try_order(row[4])
                    if ov:
                        order_no_val = ov
                if len(row) > 6 and row[6] and not deadline:
                    deadline = _try_date(row[6])

        if order_no_val:
            order_plan[order_no_val] = {
                'planned_qty':  planned_qty,
                'planned_date': deadline or '',
                'process':      product_name or '',
            }

    wb111.close()
    print(f"  -> {len(order_plan)} orders loaded from 11.1: {list(order_plan.keys())[:5]}")
except Exception as e:
    print(f"WARN: Could not load 11.1: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3：建立輸出 Excel
# ─────────────────────────────────────────────────────────────────────────────
print("STEP3: building output Excel ...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

# Sheet 01 欄位標題
HEADERS_01 = [
    "序號", "生產日期", "訂單編號", "產品品號/客戶代號",
    "設施站點", "FOSB / 批號",
    "計劃投入量", "實際投入量",
    "良品數量", "不良品數量", "良率%",
    "不良原因", "生產人員", "備註",
    "計劃完工日", "負責人(手填)", "QC簽核(手填)"
]
COL_W_01 = [6, 12, 14, 26, 10, 18, 10, 10, 10, 10, 8, 22, 8, 20, 14, 12, 12]

# ── 工作表 01：彙整總表 ────────────────────────────────────────────────────
ws1 = wb_out.create_sheet("01_彙整總表")

# 大標題
ws1.merge_cells("A1:Q1")
c = ws1["A1"]
c.value = "生產計劃暨實績彙整總表  （資料來源：11.5 生產日報_目檢合一表單）"
c.font  = Font(bold=True, size=14, color="1F4E79")
c.alignment = CENTER
ws1.row_dimensions[1].height = 30

# 表頭
for i, h in enumerate(HEADERS_01, 1):
    hcell(ws1, 2, i, h)
ws1.row_dimensions[2].height = 36

# 資料列
for idx, rec in enumerate(records, 1):
    r    = idx + 2
    plan = order_plan.get(rec['order_no'], {})

    dcell(ws1, r,  1, idx, CENTER)
    dcell(ws1, r,  2, rec['date'],                        CENTER)
    dcell(ws1, r,  3, rec['order_no'],                    CENTER)
    dcell(ws1, r,  4, rec['product'])
    dcell(ws1, r,  5, rec['station'],                     CENTER)
    dcell(ws1, r,  6, rec['fosb'])
    dcell(ws1, r,  7, plan.get('planned_qty', ''),        CENTER)
    dcell(ws1, r,  8, rec['input'],                       CENTER)
    dcell(ws1, r,  9, rec['good'],                        CENTER)
    dcell(ws1, r, 10, rec['bad'],                         CENTER)

    # 良率欄（帶色碼）
    yc = ws1.cell(row=r, column=11)
    yc.alignment = CENTER
    yc.border    = BORDER
    if rec['yield_pct'] is not None:
        yc.value         = rec['yield_pct']
        yc.number_format = "0.0%"
        if rec['yield_pct'] < 0.90:
            yc.fill = RED_FILL
        elif rec['yield_pct'] < 0.95:
            yc.fill = YLW_FILL
        else:
            yc.fill = GRN_FILL

    dcell(ws1, r, 12, rec['defect'])
    dcell(ws1, r, 13, rec['operator'],            CENTER)
    dcell(ws1, r, 14, rec['remark'])
    dcell(ws1, r, 15, plan.get('planned_date', ''), CENTER)
    dcell(ws1, r, 16, '',                          CENTER)   # 負責人 (手填)
    dcell(ws1, r, 17, '',                          CENTER)   # QC簽核 (手填)

# 欄寬 & 凍結
for i, w in enumerate(COL_W_01, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A3"

# ── 工作表 02：品質趨勢 ────────────────────────────────────────────────────
ws2 = wb_out.create_sheet("02_品質趨勢")

ws2.merge_cells("A1:H1")
ws2["A1"].value     = "品質趨勢月度摘要"
ws2["A1"].font      = Font(bold=True, size=14, color="1F4E79")
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 30

# 月度彙總計算
monthly = defaultdict(lambda: {
    'input': 0, 'good': 0, 'bad': 0, 'count': 0,
    'defect_types': defaultdict(int),
    'stations': defaultdict(int),
})

DEFECT_KEYS = ['Chipping', 'Crack', 'Scratch', 'Stain', 'Broken', 'Other']

for rec in records:
    dm = re.match(r'^(\d{4}/\d{2})', rec['date'])
    if not dm:
        continue
    ym = dm.group(1)
    m  = monthly[ym]
    m['count'] += 1
    if rec['input']:
        m['input'] += rec['input']
    if rec['good']:
        m['good'] += rec['good']
    if rec['bad']:
        m['bad'] += rec['bad']
    if rec['station']:
        m['stations'][rec['station']] += 1
    if rec['defect']:
        dl = rec['defect'].lower()
        for dk in DEFECT_KEYS:
            if dk.lower() in dl:
                m['defect_types'][dk] += 1

# 月度表頭
MONTH_HDRS = ["月份", "記錄筆數", "總投入量", "良品總數", "不良品總數", "月良率%", "主要不良類型(前3)", "製程站點分佈"]
for i, h in enumerate(MONTH_HDRS, 1):
    hcell(ws2, 2, i, h)
ws2.row_dimensions[2].height = 36

for mi, (ym, m) in enumerate(sorted(monthly.items()), 1):
    r      = mi + 2
    total  = m['good'] + m['bad']
    yld    = m['good'] / total if total > 0 else None

    top_def = sorted(m['defect_types'].items(), key=lambda x: -x[1])
    def_str = ", ".join(f"{k}({v})" for k, v in top_def[:3]) if top_def else "—"

    top_st  = sorted(m['stations'].items(), key=lambda x: -x[1])
    st_str  = " / ".join(f"{k}:{v}" for k, v in top_st[:5]) if top_st else "—"

    dcell(ws2, r, 1, ym,              CENTER)
    dcell(ws2, r, 2, m['count'],      CENTER)
    dcell(ws2, r, 3, m['input'] or '', CENTER)
    dcell(ws2, r, 4, m['good']  or '', CENTER)
    dcell(ws2, r, 5, m['bad']   or '', CENTER)

    yc2 = ws2.cell(row=r, column=6)
    yc2.alignment = CENTER; yc2.border = BORDER
    if yld is not None:
        yc2.value         = yld
        yc2.number_format = "0.0%"
        yc2.fill = RED_FILL if yld < 0.90 else (YLW_FILL if yld < 0.95 else GRN_FILL)

    dcell(ws2, r, 7, def_str)
    dcell(ws2, r, 8, st_str)

# 站點統計附表
r_off = len(monthly) + 4
ws2.merge_cells(f"A{r_off}:H{r_off}")
sc = ws2.cell(r_off, 1)
sc.value = "【製程站點總作業量（36 天合計）】"
sc.font  = Font(bold=True, color="FFFFFF")
sc.fill  = BLU_SEC
ws2.row_dimensions[r_off].height = 22

r_off += 1
station_totals = defaultdict(int)
for rec in records:
    if rec['station']:
        station_totals[rec['station']] += 1

for i, (st, cnt) in enumerate(sorted(station_totals.items(), key=lambda x: -x[1]), 1):
    dcell(ws2, r_off, (i - 1) * 2 + 1, st,  CENTER)
    dcell(ws2, r_off, (i - 1) * 2 + 2, cnt, CENTER)

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 10
ws2.column_dimensions["G"].width = 30
ws2.column_dimensions["H"].width = 32
ws2.freeze_panes = "A3"

# ── 工作表 03：新版範本 ────────────────────────────────────────────────────
ws3 = wb_out.create_sheet("03_新版範本")

ws3.merge_cells("A1:Q1")
ws3["A1"].value     = "生產計劃暨實績記錄表（改版範本 v2.0）"
ws3["A1"].font      = Font(bold=True, size=14, color="1F4E79")
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:Q2")
ws3["A2"].value = ("依據 ISO 9001:2015  §8.5.1 生產控制 / §8.7.1 不合格品 / "
                   "§9.1.1 績效評估 / §7.5.3 可追溯性")
ws3["A2"].font      = Font(size=9, color="666666", italic=True)
ws3["A2"].alignment = CENTER
ws3.row_dimensions[2].height = 18

# 基本資訊段
ws3.merge_cells("A3:Q3")
ws3["A3"].value = "【基本資訊】"
ws3["A3"].font  = Font(bold=True, color="FFFFFF")
ws3["A3"].fill  = BLU_SEC
ws3.row_dimensions[3].height = 20

info_labels = [
    ("年度",    "A4:B4"),  ("訂單編號",   "C4:D4"), ("製程/機台",   "E4:F4"),
    ("計劃量",  "G4:H4"),  ("預定完工日", "I4:J4"), ("實際完工日", "K4:L4"),
    ("負責人",  "M4:N4"),
]
for label, cellrange in info_labels:
    start = cellrange.split(":")[0]
    col_letter = re.match(r'([A-Z]+)', start).group(1)
    col_idx    = openpyxl.utils.column_index_from_string(col_letter)
    lc = ws3.cell(row=4, column=col_idx, value=label)
    lc.font = Font(bold=True); lc.fill = SUB_FILL
    lc.alignment = CENTER; lc.border = BORDER
    # Value cell (next column)
    vc = ws3.cell(row=4, column=col_idx + 1, value="")
    vc.border = BORDER; vc.alignment = LEFT
ws3.row_dimensions[4].height = 22

# 逐日記錄段
ws3.merge_cells("A5:Q5")
ws3["A5"].value = "【逐日生產實績記錄】"
ws3["A5"].font  = Font(bold=True, color="FFFFFF")
ws3["A5"].fill  = BLU_SEC
ws3.row_dimensions[5].height = 20

for i, h in enumerate(HEADERS_01, 1):
    hcell(ws3, 6, i, h)
ws3.row_dimensions[6].height = 36

for r in range(7, 25):
    for c in range(1, 18):
        cell = ws3.cell(row=r, column=c, value="")
        cell.border = BORDER
        if c == 11:
            cell.number_format = "0.0%"
    ws3.row_dimensions[r].height = 18

# 注意事項
ws3.merge_cells("A25:Q25")
ws3["A25"].value = (
    "注意事項：①良率低於 95% 請標黃；低於 90% 需填寫異常處理措施。"
    "②負責人欄必填。③QC 簽核須每日確認。④FOSB 批號需可對應至 11.5 生產日報。"
)
ws3["A25"].font      = Font(size=9, color="C00000", italic=True)
ws3["A25"].alignment = LEFT

for i, w in enumerate(COL_W_01, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A7"

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4：儲存
# ─────────────────────────────────────────────────────────────────────────────
wb_out.save(OUTPUT)
print(f"OK: Saved -> {OUTPUT}")
print(f"OK: Total records  : {len(records)}")
print(f"OK: Months covered : {sorted(monthly.keys())}")
print(f"OK: Orders matched : {sum(1 for r in records if r['order_no'] in order_plan)}/{len(records)}")
print("DONE")
