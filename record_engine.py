from __future__ import annotations

import json
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TEMPLATES = [
    {
        "code": "env_record",
        "title": "6.1 \u74b0\u5883\u76e3\u63a7\u8a18\u9304\u8868",
        "description": "\u4f9d\u76ee\u524d\u74b0\u5883\u76e3\u63a7\u8cc7\u6599\u6574\u7406\u6bcf\u65e5\u76e3\u63a7\u8a18\u9304\u3002",
        "requires": ["env_records"],
        "bundle": False,
        "downstream": ["env_monthly_pack"],
        "includes": [],
    },
    {
        "code": "env_monthly_pack",
        "title": "6 \u5de5\u4f5c\u74b0\u5883\u76e3\u63a7\u6708\u5831\u5305",
        "description": "\u4e00\u6b21\u7522\u751f\u74b0\u5883\u76e3\u63a7\u6708\u5831\u6458\u8981\u8207 6.1 \u74b0\u5883\u76e3\u63a7\u8a18\u9304\u8868\u660e\u7d30\u3002",
        "requires": ["env_records"],
        "bundle": True,
        "downstream": [],
        "includes": ["env_record"],
    },
    {
        "code": "production_daily",
        "title": "11.5 \u751f\u7522\u65e5\u5831\u8868",
        "description": "\u5c07\u76ee\u524d\u751f\u7522\u6279\u6b21\u6574\u7406\u6210\u751f\u7522\u65e5\u5831\u3002",
        "requires": ["prod_records"],
        "bundle": False,
        "downstream": ["shipment_order_143", "shipping_pack"],
        "includes": [],
    },
    {
        "code": "quality_incoming",
        "title": "12.1 \u54c1\u8cea\u7ba1\u7406\u8a18\u9304\u8868(\u539f\u6599)",
        "description": "\u5c07\u9032\u6599\u6aa2\u9a57\u8cc7\u6599\u6574\u7406\u6210\u539f\u6599\u54c1\u8cea\u8a18\u9304\u3002",
        "requires": ["quality_records"],
        "bundle": False,
        "downstream": ["shipping_inspection_125"],
        "includes": [],
    },
    {
        "code": "material_request_112",
        "title": "11.2 \u9818\u6599\u55ae",
        "description": "\u4f9d\u51fa\u8ca8\u6216\u751f\u7522\u9700\u6c42\u7522\u751f\u9818\u6599\u55ae\u8349\u7a3f\u3002",
        "requires": ["shipment_request"],
        "bundle": False,
        "downstream": [],
        "includes": [],
    },
    {
        "code": "shipment_order_143",
        "title": "14.3 \u51fa\u8ca8\u55ae",
        "description": "\u4f9d\u8a02\u55ae\u8207 LOT \u7522\u751f\u51fa\u8ca8\u55ae\u8349\u7a3f\u3002",
        "requires": ["shipment_request"],
        "bundle": False,
        "downstream": ["shipping_pack"],
        "includes": [],
    },
    {
        "code": "shipping_inspection_125",
        "title": "12.5 \u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304\u8868",
        "description": "\u4f9d\u8a02\u55ae\u8207\u54c1\u8cea\u8cc7\u6599\u6574\u7406\u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304\u3002",
        "requires": ["shipment_request"],
        "bundle": False,
        "downstream": ["shipping_pack"],
        "includes": [],
    },
    {
        "code": "shipping_check_145",
        "title": "14.5 \u51fa\u8ca8\u6aa2\u67e5\u8868",
        "description": "\u4f9d\u51fa\u8ca8\u8cc7\u6599\u7522\u751f\u51fa\u8ca8\u524d\u6aa2\u67e5\u8868\u3002",
        "requires": ["shipment_request"],
        "bundle": False,
        "downstream": ["shipping_pack"],
        "includes": [],
    },
    {
        "code": "shipping_pack",
        "title": "\u51fa\u8ca8\u6d41\u7a0b\u7d00\u9304\u5305",
        "description": "\u4e00\u6b21\u7522\u751f\u51fa\u8ca8\u55ae\u3001\u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304\u8207\u51fa\u8ca8\u6aa2\u67e5\u8868\u3002",
        "requires": ["shipment_request"],
        "bundle": True,
        "downstream": [],
        "includes": ["shipment_order_143", "shipping_inspection_125", "shipping_check_145"],
    },
    {
        "code": "cip_152",
        "title": "15.2 \u88fd\u7a0b\u7f3a\u9677\u8ffd\u8e64\u6539\u5584(CIP)\u7d00\u9304\u8868",
        "description": "\u4f9d\u4e0d\u7b26\u5408\u8cc7\u6599\u6574\u7406\u6210 CIP \u6539\u5584\u8ffd\u8e64\u8868\u3002",
        "requires": ["nonconformance"],
        "bundle": False,
        "downstream": ["cip_pack"],
        "includes": [],
    },
    {
        "code": "cip_pack",
        "title": "CIP / \u4e0d\u7b26\u5408\u6d41\u7a0b\u5305",
        "description": "\u4e00\u6b21\u7522\u751f CIP \u7d00\u9304\u8868\u8207\u4e0d\u7b26\u5408\u4f86\u6e90\u6458\u8981\u3002",
        "requires": ["nonconformance"],
        "bundle": True,
        "downstream": [],
        "includes": ["cip_152"],
    },
    {
        "code": "audit_notice",
        "title": "9 內部稽核通知書",
        "description": "依選定的稽核計畫產生內部稽核通知書與查檢範圍工作表。",
        "requires": ["audit_plans"],
        "bundle": False,
        "downstream": ["audit_pack"],
        "includes": [],
    },
    {
        "code": "audit_pack",
        "title": "稽核流程包",
        "description": "一次產生稽核通知書（含查檢範圍）ZIP 包，適合稽核前整理使用。",
        "requires": ["audit_plans"],
        "bundle": True,
        "downstream": [],
        "includes": ["audit_notice"],
    },
    {
        "code": "management_review_pack",
        "title": "16 管理審查流程包",
        "description": "自動彙整各類記錄，產生 ISO 9001 第 9.3 條所需的管理審查輸入摘要與決議記錄表。",
        "requires": [],
        "bundle": True,
        "downstream": [],
        "includes": [],
    },
]

KEYWORDS = {
    "shipping_pack": ["\u51fa\u8ca8\u6d41\u7a0b", "\u51fa\u8ca8\u6d41\u7a0b\u5305", "\u51fa\u8ca8\u5305", "\u51fa\u8ca8\u7d00\u9304\u5305", "\u51fa\u8ca8"],
    "shipment_order_143": ["\u51fa\u8ca8\u55ae", "\u51fa\u8ca8"],
    "shipping_inspection_125": ["\u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304", "\u51fa\u8ca8\u6aa2\u67e5", "\u6aa2\u67e5\u7d00\u9304"],
    "shipping_check_145": ["\u51fa\u8ca8\u6aa2\u67e5\u8868", "\u6aa2\u67e5\u8868"],
    "material_request_112": ["\u9818\u6599", "\u9818\u6599\u55ae", "\u5099\u6599"],
    "cip_152": ["\u4e0d\u7b26\u5408", "\u6539\u5584", "cip", "\u77ef\u6b63", "\u88fd\u7a0b\u7f3a\u9677\u8ffd\u8e64", "\u7570\u5e38\u6539\u5584"],
    "cip_pack": ["cip \u6d41\u7a0b\u5305", "\u4e0d\u7b26\u5408\u6d41\u7a0b\u5305", "cip \u5305", "\u6539\u5584\u6d41\u7a0b\u5305", "\u4e0d\u7b26\u5408\u5305"],
    "production_daily": ["\u751f\u7522\u65e5\u5831", "\u751f\u7522", "\u88fd\u7a0b", "\u65e5\u5831"],
    "quality_incoming": ["\u9032\u6599", "\u54c1\u8cea", "\u6aa2\u9a57", "\u539f\u6599"],
    "env_record": ["\u74b0\u5883", "\u76e3\u63a7", "\u5de5\u4f5c\u74b0\u5883", "\u7c92\u5b50"],
    "env_monthly_pack": ["\u74b0\u5883\u6708\u5831", "\u76e3\u63a7\u6708\u5831", "\u5de5\u4f5c\u74b0\u5883\u6708\u5831", "\u74b0\u5883\u76e3\u63a7\u6708\u5831", "\u7c92\u5b50\u6708\u5831"],
    "audit_notice": ["稽核通知", "內部稽核通知", "稽核通知書", "稽核計畫通知"],
    "audit_pack": ["稽核流程包", "稽核包", "稽核文件包", "內部稽核流程"],
    "management_review_pack": ["管理審查", "管理評審", "管審", "管理審查流程包", "management review", "管理審查記錄", "管理審查報告"],
}


def _tmp_path(suffix: str) -> Path:
    handle = tempfile.NamedTemporaryFile(suffix=suffix, delete=False, dir=tempfile.gettempdir())
    handle.close()
    return Path(handle.name)


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _template_index() -> dict:
    return {item["code"]: item for item in TEMPLATES}


def get_catalog() -> list[dict]:
    template_map = _template_index()
    return [
        {
            **item,
            "included_templates": [template_map[code]["title"] for code in item.get("includes", []) if code in template_map],
            "downstream_templates": [template_map[code]["title"] for code in item.get("downstream", []) if code in template_map],
        }
        for item in TEMPLATES
    ]


def suggest_templates(prompt: str, context: dict | None = None) -> list[dict]:
    text = _normalize_text(prompt).lower()
    ranked = []
    for template in TEMPLATES:
        score = 0
        for keyword in KEYWORDS.get(template["code"], []):
            if keyword.lower() in text:
                score += 3
        if template["code"] == "shipping_pack" and "\u51fa\u8ca8\u6d41\u7a0b" in prompt:
            score += 10
        if context:
            if template["code"] == "env_record" and (context.get("env_count") or 0) > 0:
                score += 1
            if template["code"] == "env_monthly_pack" and (context.get("env_count") or 0) > 0:
                score += 2
            if template["code"] == "production_daily" and (context.get("prod_count") or 0) > 0:
                score += 1
            if template["code"] == "quality_incoming" and (context.get("quality_count") or 0) > 0:
                score += 1
            if template["requires"] == ["shipment_request"] and (context.get("shipment_order_count") or 0) > 0:
                score += 1
            if template["requires"] == ["nonconformance"] and (context.get("nonconformance_count") or 0) > 0:
                score += 1
            if template["code"] == "management_review_pack":
                data_sources = (
                    (1 if (context.get("prod_count") or 0) > 0 else 0) +
                    (1 if (context.get("quality_count") or 0) > 0 else 0) +
                    (1 if (context.get("env_count") or 0) > 0 else 0) +
                    (1 if (context.get("nonconformance_count") or 0) > 0 else 0)
                )
                score += data_sources
        if score > 0:
            ranked.append((score, template))
    ranked.sort(key=lambda item: (-item[0], item[1]["title"]))
    return [item[1] for item in ranked[:5]]


def _missing_detail(scope: str, field_key: str, label: str, row_index: int | None = None, item_id: str | None = None, scope_label: str | None = None) -> dict:
    result = {"scope": scope, "field_key": field_key, "label": label}
    if row_index is not None:
        result["row_index"] = row_index
    if item_id:
        result["item_id"] = item_id
    if scope_label:
        result["scope_label"] = scope_label
    return result


def _shipment_missing_details(payload: dict) -> list[dict]:
    request = payload.get("shipment_request") or {}
    mapping = [
        ("date", "\u51fa\u8ca8\u65e5\u671f"),
        ("department", "\u7533\u8acb\u90e8\u9580"),
        ("requester", "\u7533\u8acb\u4eba"),
        ("product_name", "\u7522\u54c1\u540d\u7a31"),
        ("quantity", "\u6578\u91cf"),
    ]
    details = []
    for field_key, label in mapping:
        if not _normalize_text(request.get(field_key)):
            details.append(_missing_detail("shipment", field_key, label, scope_label="\u51fa\u8ca8\u55ae\u8349\u7a3f"))
    return details


def _nonconformance_required_details(payload: dict, template_code: str) -> list[dict]:
    item = payload.get("nonconformance") or {}
    mapping = [
        ("id", "\u4e0d\u7b26\u5408\u7de8\u865f"),
        ("date", "\u767c\u751f\u65e5"),
        ("dept", "\u90e8\u9580"),
        ("description", "\u554f\u984c\u63cf\u8ff0"),
        ("responsible", "\u8cac\u4efb\u4eba"),
    ]
    item_id = _normalize_text(item.get("id"))
    details = []
    for field_key, label in mapping:
        if not _normalize_text(item.get(field_key)):
            details.append(_missing_detail("nonconformance", field_key, label, item_id=item_id, scope_label="\u4e0d\u7b26\u5408\u7ba1\u7406"))
    return details


def _collect_nonconformance_warnings(payload: dict, template_code: str) -> list[str]:
    item = payload.get("nonconformance") or {}
    warnings = []
    optional_labels = [
        ("severity", "\u56b4\u91cd\u5ea6"),
        ("rootCause", "\u539f\u56e0\u5206\u6790"),
        ("correctiveAction", "\u6539\u5584\u63aa\u65bd"),
        ("dueDate", "\u5230\u671f\u65e5"),
        ("status", "\u72c0\u614b"),
    ]
    missing_optional = [label for field, label in optional_labels if not _normalize_text(item.get(field))]
    if missing_optional:
        warnings.append("CIP \u8349\u7a3f\u4ecd\u5efa\u8b70\u88dc\u9f4a\uff1a" + " / ".join(missing_optional))
    if not _normalize_text(item.get("effectiveness")):
        warnings.append("\u5efa\u8b70\u88dc\u4e0a\u6548\u679c\u78ba\u8a8d\uff0c\u65b9\u4fbf\u5f8c\u7e8c\u8ffd\u8e64\u6539\u5584\u662f\u5426\u6709\u6548\u3002")
    status_text = _normalize_text(item.get("status"))
    if status_text in {"\u5df2\u7d50\u6848", "Closed"} and not _normalize_text(item.get("closeDate")):
        warnings.append("\u6b64\u7b46\u4e0d\u7b26\u5408\u5df2\u6a19\u793a\u7d50\u6848\uff0c\u5efa\u8b70\u88dc\u4e0a\u7d50\u6848\u65e5\u671f\u3002")
    if template_code == "cip_pack" and not _normalize_text(item.get("closeDate")) and status_text not in {"\u5df2\u7d50\u6848", "Closed"}:
        warnings.append("CIP \u6d41\u7a0b\u5305\u5efa\u8b70\u88dc\u4e0a\u9810\u8a08\u7d50\u6848\u6642\u9593\u6216\u9032\u5ea6\u8aaa\u660e\uff0c\u65b9\u4fbf\u5f8c\u7e8c\u8ffd\u8e64\u3002")
    return warnings


def _nonconformance_item(payload: dict) -> dict:
    item = payload.get("nonconformance") or {}
    return {
        "id": _normalize_text(item.get("id")),
        "date": _normalize_text(item.get("date")),
        "dept": _normalize_text(item.get("dept")),
        "type": _normalize_text(item.get("type")) or "\u88fd\u7a0b\u7570\u5e38",
        "severity": _normalize_text(item.get("severity")) or "\u672a\u8a55\u4f30",
        "description": _normalize_text(item.get("description")),
        "rootCause": _normalize_text(item.get("rootCause")),
        "correctiveAction": _normalize_text(item.get("correctiveAction")),
        "responsible": _normalize_text(item.get("responsible")),
        "dueDate": _normalize_text(item.get("dueDate")),
        "status": _normalize_text(item.get("status")) or "\u5f85\u8655\u7406",
        "closeDate": _normalize_text(item.get("closeDate")),
        "effectiveness": _normalize_text(item.get("effectiveness")),
        "source_file": _normalize_text(item.get("source_file")),
    }


def _nonconformance_next_action(item: dict) -> str:
    status = _normalize_text(item.get("status"))
    if status in {"\u5df2\u7d50\u6848", "Closed"}:
        if _normalize_text(item.get("effectiveness")):
            return "\u5efa\u8b70\u4fdd\u7559\u6548\u679c\u78ba\u8a8d\u8b49\u64da\uff0c\u4f5c\u70ba\u7d50\u6848\u5f8c\u8ffd\u8e64\u8cc7\u6599\u3002"
        return "\u5df2\u6a19\u793a\u7d50\u6848\uff0c\u5efa\u8b70\u88dc\u4e0a\u6548\u679c\u78ba\u8a8d\u8aaa\u660e\u3002"
    if _normalize_text(item.get("correctiveAction")) and _normalize_text(item.get("dueDate")):
        return "\u53ef\u4f9d\u73fe\u6709\u6539\u5584\u63aa\u65bd\u8207\u5230\u671f\u65e5\u7e7c\u7e8c\u8ffd\u8e64\u57f7\u884c\u6210\u6548\u3002"
    return "\u8acb\u5148\u88dc\u9f4a\u6539\u5584\u63aa\u65bd\u3001\u8cac\u4efb\u4eba\u8207\u5230\u671f\u6642\u9593\uff0c\u518d\u9032\u884c CIP \u8ffd\u8e64\u3002"


def precheck_template(payload: dict) -> dict:
    template_code = _normalize_text(payload.get("template_code"))
    if not template_code:
        raise ValueError("template_code is required.")
    template_map = _template_index()
    template = template_map.get(template_code)
    if not template:
        raise ValueError(f"Unsupported template_code: {template_code}")

    missing_details = []
    warnings = []
    if "env_records" in template["requires"] and not (payload.get("env_records") or []):
        missing_details.append(_missing_detail("environment", "records", "\u74b0\u5883\u76e3\u63a7\u8cc7\u6599", scope_label="\u74b0\u5883\u76e3\u63a7"))
    if "prod_records" in template["requires"] and not (payload.get("prod_records") or []):
        missing_details.append(_missing_detail("production", "records", "\u751f\u7522\u6279\u6b21\u8cc7\u6599", scope_label="\u751f\u7522\u6279\u6b21"))
    if "quality_records" in template["requires"] and not (payload.get("quality_records") or []):
        missing_details.append(_missing_detail("quality", "records", "\u9032\u6599\u6aa2\u9a57\u8cc7\u6599", scope_label="\u9032\u6599\u6aa2\u9a57"))
    if "shipment_request" in template["requires"]:
        missing_details.extend(_shipment_missing_details(payload))
        request = payload.get("shipment_request") or {}
        if template_code == "shipping_pack":
            if not _normalize_text(request.get("spec")):
                warnings.append("\u51fa\u8ca8\u6d41\u7a0b\u5305\u5efa\u8b70\u88dc\u4e0a\u898f\u683c\uff0c\u65b9\u4fbf\u4e0b\u6e38\u8868\u55ae\u5b8c\u6574\u5e36\u503c\u3002")
            if not _normalize_text(request.get("unit")):
                warnings.append("\u51fa\u8ca8\u6d41\u7a0b\u5305\u5efa\u8b70\u88dc\u4e0a\u55ae\u4f4d\uff0c\u907f\u514d\u4e0b\u6e38\u8868\u55ae\u4ecd\u9700\u624b\u52d5\u88dc\u503c\u3002")
    if "nonconformance" in template["requires"]:
        missing_details.extend(_nonconformance_required_details(payload, template_code))
        warnings.extend(_collect_nonconformance_warnings(payload, template_code))
    if "audit_plans" in template["requires"]:
        missing_details.extend(_audit_required_details(payload))

    if template_code == "shipping_pack":
        warnings.append("\u51fa\u8ca8\u6d41\u7a0b\u5305\u6703\u4e00\u6b21\u7522\u751f\u51fa\u8ca8\u55ae\u3001\u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304\u8207\u51fa\u8ca8\u6aa2\u67e5\u8868\u3002")
    if template_code == "env_monthly_pack":
        warnings.append("\u74b0\u5883\u76e3\u63a7\u6708\u5831\u5305\u6703\u540c\u6642\u5305\u542b\u6708\u5831\u6458\u8981\u8207 6.1 \u74b0\u5883\u76e3\u63a7\u8a18\u9304\u660e\u7d30\u3002")
    if template_code == "cip_152":
        warnings.append("CIP \u8349\u7a3f\u53ef\u5148\u7522\u751f\uff0c\u518d\u56de\u5230\u4e0d\u7b26\u5408\u7ba1\u7406\u88dc\u5b8c\u539f\u56e0\u5206\u6790\u8207\u6539\u5584\u5167\u5bb9\u3002")
    if template_code == "cip_pack":
        warnings.append("CIP / \u4e0d\u7b26\u5408\u6d41\u7a0b\u5305\u6703\u5305\u542b\u4e00\u4efd CIP \u7d00\u9304\u8868\u8207\u4e00\u4efd\u4e0d\u7b26\u5408\u4f86\u6e90\u6458\u8981\u3002")
    if template_code == "audit_notice":
        warnings.append("稽核通知書包含通知書本體與稽核查檢範圍工作表，請稽核員於執行稽核時填入查驗結果。")
    if template_code == "audit_pack":
        warnings.append("稽核流程包含稽核通知書（含查檢範圍），ZIP 解壓後可直接使用。")
    if template_code == "management_review_pack":
        warnings.append("管理審查流程包包含：輸入摘要（自動帶入現有資料）與決議記錄表（空白模板）。")
        if not any([
            payload.get("prod_records"), payload.get("quality_records"),
            payload.get("env_records"), payload.get("audit_plans"),
            payload.get("all_nonconformances"),
        ]):
            warnings.append("目前各類記錄均為空，建議先上傳各類記錄後再產生，以獲得完整摘要。")

    downstream_templates = [
        {"code": template_map[code]["code"], "title": template_map[code]["title"], "description": template_map[code]["description"]}
        for code in template.get("downstream", []) if code in template_map
    ]
    included_templates = [
        {"code": template_map[code]["code"], "title": template_map[code]["title"]}
        for code in template.get("includes", []) if code in template_map
    ]

    missing_items = []
    for detail in missing_details:
        label = detail.get("label") or detail.get("scope_label") or detail.get("scope") or "\u7f3a\u5c11\u6b04\u4f4d"
        if label not in missing_items:
            missing_items.append(label)

    return {
        "template_code": template_code,
        "title": template["title"],
        "bundle": template["bundle"],
        "ready": len(missing_details) == 0,
        "missing_items": missing_items,
        "missing_details": missing_details,
        "warnings": warnings,
        "source_counts": {
            "env_records": len(payload.get("env_records") or []),
            "prod_records": len(payload.get("prod_records") or []),
            "quality_records": len(payload.get("quality_records") or []),
            "shipment": 1 if payload.get("shipment_request") else 0,
            "nonconformance": 1 if payload.get("nonconformance") else 0,
            "audit_plans": len(payload.get("audit_plans") or []),
        },
        "included_templates": included_templates,
        "downstream_templates": downstream_templates,
    }


def _new_wb(title: str) -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def _tmp_xlsx() -> Path:
    return _tmp_path('.xlsx')


# ── Excel 格式化輔助 ──────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _hdr_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _style_header_row(ws, row: int = 1, fill_hex: str = "1E3A5F",
                       font_color: str = "FFFFFF", bold: bool = True) -> None:
    """把指定列設為深色底、白色粗體標題。"""
    fill = _hdr_fill(fill_hex)
    font = Font(bold=bold, color=font_color, name="Calibri", size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

def _style_data_rows(ws, start_row: int = 2) -> None:
    """把資料列加細框線，奇偶列輕微底色。"""
    even_fill = _hdr_fill("F5F7FA")
    odd_fill  = _hdr_fill("FFFFFF")
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=1):
        for cell in row:
            cell.fill = even_fill if i % 2 == 0 else odd_fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")

def _style_row_by_result(ws, row_idx: int, result: str) -> None:
    """依結果欄標紅/黃底色（不合格=紅，警告=黃）。"""
    result_norm = (result or "").strip()
    if result_norm == "不合格":
        fill = _hdr_fill("FFDEDE")
        font_color = "CC0000"
    elif result_norm == "警告":
        fill = _hdr_fill("FFF5CC")
        font_color = "886600"
    else:
        return
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = Font(color=font_color, name="Calibri", size=10)

def _autofit_columns(ws, min_width: int = 8, max_width: int = 40) -> None:
    """自動調整欄寬（取最長字元數估算）。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                # 中文字元寬度約 2 倍
                display_len = sum(2 if ord(c) > 0x7F else 1 for c in val)
                max_len = max(max_len, display_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _parse_record_datetime(item: dict) -> datetime | None:
    raw = _normalize_text(item.get("dateTime") or item.get("date"))
    if not raw:
        return None
    candidates = ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"]
    for pattern in candidates:
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            continue
    return None


def _sort_env_records(records: list[dict]) -> list[dict]:
    def key(item: dict):
        parsed = _parse_record_datetime(item)
        return (
            parsed or datetime.max,
            _normalize_text(item.get("point")),
            _normalize_text(item.get("location")),
        )

    return sorted(records, key=key)


def _env_pack_period(records: list[dict]) -> tuple[str, str]:
    parsed_dates = [dt for dt in (_parse_record_datetime(item) for item in records) if dt]
    if not parsed_dates:
        return "unknown", "\u672a\u77e5\u671f\u9593"
    start = min(parsed_dates)
    end = max(parsed_dates)
    start_tag = start.strftime("%Y%m")
    end_tag = end.strftime("%Y%m")
    label = start.strftime("%Y/%m/%d") if start.date() == end.date() else f"{start.strftime('%Y/%m/%d')} ~ {end.strftime('%Y/%m/%d')}"
    return (start_tag if start_tag == end_tag else f"{start_tag}-{end_tag}", label)


def _build_env_record(payload: dict):
    wb, ws = _new_wb("環境監控記錄")
    headers = ["日期", "點位", "地點", "0.3μm", "0.5μm", "5.0μm", "溫度", "濕度", "正壓", "記錄者", "結果"]
    ws.append(headers)
    sorted_records = _sort_env_records(list(payload.get("env_records") or []))
    for r_idx, item in enumerate(sorted_records, start=2):
        result = _normalize_text(item.get("result")) or ""
        ws.append([
            item.get("date"), item.get("point"), item.get("location"),
            item.get("particles03"), item.get("particles05"),
            item.get("particles5") or item.get("particles50"),
            item.get("temp"), item.get("humidity"), item.get("pressure"),
            item.get("operator"), item.get("result"),
        ])
        _style_row_by_result(ws, r_idx, result)
    _style_header_row(ws, row=1, fill_hex="0F4C81")
    _style_data_rows(ws, start_row=2)
    # 異常列覆蓋底色（style_data_rows 會清掉，再補回來）
    for r_idx, item in enumerate(sorted_records, start=2):
        result = _normalize_text(item.get("result")) or ""
        _style_row_by_result(ws, r_idx, result)
    _autofit_columns(ws)
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "6.1_環境監控記錄表.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_env_monthly_summary(payload: dict):
    records = _sort_env_records(list(payload.get("env_records") or []))
    period_tag, period_label = _env_pack_period(records)
    total = len(records)
    passed = sum(1 for item in records if _normalize_text(item.get("result")) == "合格")
    warnings = sum(1 for item in records if _normalize_text(item.get("result")) == "警告")
    failed = sum(1 for item in records if _normalize_text(item.get("result")) == "不合格")

    pass_rate = round((passed / total) * 100, 1) if total else 0
    anomaly_count = warnings + failed

    wb, ws = _new_wb("月報摘要")

    # ── 封面（插入為第一張工作表）──────────────────────
    cover = wb.create_sheet("封面", 0)
    # 公司名稱大標題
    cover["A1"] = "潔沛企業有限公司"
    cover["A1"].font = Font(bold=True, size=18, color="1E3A5F", name="Calibri")
    cover["A2"] = "6 工作環境監控月報"
    cover["A2"].font = Font(bold=True, size=14, color="2563EB", name="Calibri")
    cover["A3"] = f"統計期間：{period_label}"
    cover["A3"].font = Font(size=11, color="475569", name="Calibri")

    # KPI 小表（A5:B9）
    kpi_rows = [
        ("項目", "數值"),
        ("量測總筆數", total),
        ("合格筆數", passed),
        ("警告筆數", warnings),
        ("不合格筆數", failed),
        ("合格率", f"{pass_rate}%"),
    ]
    for r_idx, (label, val) in enumerate(kpi_rows, start=5):
        cover.cell(r_idx, 1, label)
        cover.cell(r_idx, 2, val)
    # 標題列格式
    _style_header_row(cover, row=5, fill_hex="1E3A5F")
    # 合格率列特殊顏色
    pass_fill = _hdr_fill("D1FAE5") if pass_rate >= 95 else _hdr_fill("FEF3C7") if pass_rate >= 85 else _hdr_fill("FEE2E2")
    pass_font_color = "065F46" if pass_rate >= 95 else "92400E" if pass_rate >= 85 else "991B1B"
    for col in [1, 2]:
        cell = cover.cell(10, col)
        cell.fill = pass_fill
        cell.font = Font(bold=True, color=pass_font_color, name="Calibri", size=11)
        cell.border = _BORDER
    # 資料列框線
    for r_idx in range(6, 11):
        for col in [1, 2]:
            cell = cover.cell(r_idx, col)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")

    if anomaly_count == 0:
        summary_text = "✅ 本月所有量測記錄均合格，無異常。"
    else:
        summary_text = (
            f"⚠ 本月共 {anomaly_count} 筆異常"
            f"（警告 {warnings} 筆、不合格 {failed} 筆），"
            "請參閱「異常明細」工作表。"
        )
    cover["A12"] = "摘要說明"
    cover["A12"].font = Font(bold=True, color="1E3A5F", name="Calibri")
    cover["B12"] = summary_text
    cover["B12"].font = Font(
        color="065F46" if anomaly_count == 0 else "991B1B",
        name="Calibri"
    )
    cover["A14"] = "製表人"
    cover["B14"] = ""
    cover["A15"] = "主管核准"
    cover["B15"] = ""
    cover["A16"] = "製表日期"
    cover["B16"] = datetime.today().strftime("%Y-%m-%d")
    for row_n in [14, 15, 16]:
        for col in [1, 2]:
            cover.cell(row_n, col).border = _BORDER
    cover.column_dimensions["A"].width = 16
    cover.column_dimensions["B"].width = 36

    # ── 月報摘要工作表 ─────────────────────────────────
    ws["A1"] = "6 工作環境監控月報摘要"
    ws["A1"].font = Font(bold=True, size=13, color="1E3A5F", name="Calibri")
    summary_data = [
        ("項目", "數值"),
        ("統計期間", period_label),
        ("總筆數", total),
        ("合格", passed),
        ("警告", warnings),
        ("不合格", failed),
        ("合格率", f"{pass_rate}%"),
    ]
    for r_idx, (label, val) in enumerate(summary_data, start=2):
        ws.cell(r_idx, 1, label)
        ws.cell(r_idx, 2, val)
    _style_header_row(ws, row=2, fill_hex="1E3A5F")
    for r_idx in range(3, 2 + len(summary_data)):
        for col in [1, 2]:
            ws.cell(r_idx, col).border = _BORDER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 36

    # ── 日別統計工作表 ─────────────────────────────────
    by_day = wb.create_sheet("日別統計")
    by_day.append(["日期", "總筆數", "合格", "警告", "不合格"])
    day_stats: dict[str, dict[str, int]] = {}
    for item in records:
        day = _normalize_text(item.get("date")) or "未填日期"
        stat = day_stats.setdefault(day, {"total": 0, "pass": 0, "warn": 0, "fail": 0})
        stat["total"] += 1
        result = _normalize_text(item.get("result"))
        if result == "合格":
            stat["pass"] += 1
        elif result == "警告":
            stat["warn"] += 1
        elif result == "不合格":
            stat["fail"] += 1
    for day in sorted(day_stats):
        stat = day_stats[day]
        by_day.append([day, stat["total"], stat["pass"], stat["warn"], stat["fail"]])
    _style_header_row(by_day, row=1, fill_hex="1E4E8C")
    _style_data_rows(by_day, start_row=2)
    # 有警告/不合格的列標色
    for r_idx, day in enumerate(sorted(day_stats), start=2):
        stat = day_stats[day]
        if stat["fail"] > 0:
            for col in range(1, 6):
                by_day.cell(r_idx, col).fill = _hdr_fill("FFDEDE")
        elif stat["warn"] > 0:
            for col in range(1, 6):
                by_day.cell(r_idx, col).fill = _hdr_fill("FFF5CC")
    _autofit_columns(by_day)

    # ── 點位統計工作表 ─────────────────────────────────
    by_point = wb.create_sheet("點位統計")
    by_point.append(["點位", "地點", "總筆數", "警告", "不合格", "0.3μm最大值", "0.5μm最大值", "5.0μm最大值"])
    point_stats: dict[tuple[str, str], dict[str, float]] = {}
    for item in records:
        point = _normalize_text(item.get("point")) or "未填點位"
        location = _normalize_text(item.get("location"))
        key = (point, location)
        stat = point_stats.setdefault(key, {"total": 0, "warn": 0, "fail": 0, "p03": 0, "p05": 0, "p5": 0})
        stat["total"] += 1
        result = _normalize_text(item.get("result"))
        if result == "警告":
            stat["warn"] += 1
        elif result == "不合格":
            stat["fail"] += 1
        stat["p03"] = max(stat["p03"], float(item.get("particles03") or 0))
        stat["p05"] = max(stat["p05"], float(item.get("particles05") or 0))
        stat["p5"] = max(stat["p5"], float(item.get("particles5") or item.get("particles50") or 0))
    for (point, location), stat in sorted(point_stats.items(), key=lambda pair: (pair[0][0], pair[0][1])):
        by_point.append([point, location, int(stat["total"]), int(stat["warn"]), int(stat["fail"]), stat["p03"], stat["p05"], stat["p5"]])
    _style_header_row(by_point, row=1, fill_hex="1E4E8C")
    _style_data_rows(by_point, start_row=2)
    # 有異常的點位標色
    for r_idx, ((point, location), stat) in enumerate(
        sorted(point_stats.items(), key=lambda pair: (pair[0][0], pair[0][1])), start=2
    ):
        if stat["fail"] > 0:
            for col in range(1, 9):
                by_point.cell(r_idx, col).fill = _hdr_fill("FFDEDE")
        elif stat["warn"] > 0:
            for col in range(1, 9):
                by_point.cell(r_idx, col).fill = _hdr_fill("FFF5CC")
    _autofit_columns(by_point)

    # ── 異常明細工作表（僅列警告 / 不合格記錄）──────────
    anomaly_ws = wb.create_sheet("異常明細")
    anomaly_ws.append(["日期", "點位", "地點", "0.3μm", "0.5μm", "5.0μm", "溫度", "濕度", "正壓", "記錄者", "結果", "建議追蹤"])
    anomaly_records = [
        r for r in records
        if _normalize_text(r.get("result")) in ("警告", "不合格")
    ]
    if anomaly_records:
        for r_idx, item in enumerate(anomaly_records, start=2):
            result_val = _normalize_text(item.get("result")) or ""
            suggestion = (
                "請立即停止相關作業，確認設備狀況，調查超標原因並提出改善措施。"
                if result_val == "不合格"
                else "請加強該點位監控頻率，確認是否需要清潔或校正設備。"
            )
            anomaly_ws.append([
                item.get("date"), item.get("point"), item.get("location"),
                item.get("particles03"), item.get("particles05"),
                item.get("particles5") or item.get("particles50"),
                item.get("temp"), item.get("humidity"), item.get("pressure"),
                item.get("operator"), item.get("result"), suggestion,
            ])
            _style_row_by_result(anomaly_ws, r_idx, _normalize_text(item.get("result")) or "")
    else:
        anomaly_ws.append(["本月無異常記錄"])
    _style_header_row(anomaly_ws, row=1, fill_hex="7F1D1D" if failed > 0 else "92400E" if warnings > 0 else "1E4E8C")
    _autofit_columns(anomaly_ws)

    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, f"6_工作環境監控月報_{period_tag}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_production_daily(payload: dict):
    wb, ws = _new_wb("\u751f\u7522\u65e5\u5831")
    ws.append(["\u6279\u865f", "\u5ba2\u6236", "\u7522\u54c1", "\u6295\u5165\u6578", "\u826f\u54c1\u6578", "\u4e0d\u826f\u6578", "\u826f\u7387", "\u4e0d\u826f\u539f\u56e0", "\u4f5c\u696d\u54e1", "\u65e5\u671f"])
    for row in payload.get("prod_records") or []:
        reasons = row.get("defectReasons") or []
        if not isinstance(reasons, list):
            reasons = [str(reasons)] if reasons else []
        ws.append([row.get("lot"), row.get("customer"), row.get("product"), row.get("input"), row.get("good"), row.get("defect"), row.get("yieldRate"), ", ".join(reasons), row.get("operator"), row.get("date")])
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "11.5_\u751f\u7522\u65e5\u5831\u8868.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_quality_incoming(payload: dict):
    wb, ws = _new_wb("\u539f\u6599\u54c1\u8cea\u8a18\u9304")
    ws.append(["\u6750\u6599\u540d\u7a31", "\u6279\u865f", "\u6578\u91cf", "\u898f\u683c", "\u6aa2\u9a57\u6578\u91cf", "PH", "\u6bd4\u91cd", "RI", "\u65cb\u5149\u503c", "\u7d50\u679c", "\u5099\u8a3b"])
    for row in payload.get("quality_records") or []:
        ws.append([row.get("materialName"), row.get("batchNo"), row.get("quantity"), row.get("spec"), row.get("inspQty"), row.get("ph"), row.get("density"), row.get("ri"), row.get("rotation"), row.get("result"), row.get("note")])
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "12.1_\u54c1\u8cea\u7ba1\u7406\u8a18\u9304\u8868(\u539f\u6599).xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_material_request(payload: dict):
    request = payload.get("shipment_request") or {}
    wb, ws = _new_wb("\u9818\u6599\u55ae")
    ws["A1"] = "11.2 \u9818\u6599\u55ae"
    ws["B3"] = request.get("department") or "\u8cc7\u6750\u8ab2"
    ws["D3"] = request.get("requester") or ""
    ws["E3"] = request.get("date") or ""
    ws["B5"] = request.get("batch_display") or request.get("order_no") or ""
    ws["C5"] = request.get("product_name") or ""
    ws["D5"] = request.get("spec") or ""
    ws["E5"] = request.get("quantity") or ""
    ws["F5"] = request.get("unit") or ""
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "11.2_\u9818\u6599\u55ae.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_shipment_order_143(payload: dict):
    request = payload.get("shipment_request") or {}
    wb, ws = _new_wb("\u51fa\u8ca8\u55ae")
    ws["A1"] = "14.3 \u51fa\u8ca8\u55ae"
    ws["B3"] = request.get("department") or ""
    ws["D3"] = request.get("requester") or ""
    ws["E3"] = request.get("date") or ""
    ws["B5"] = request.get("batch_display") or request.get("order_no") or ""
    ws["C5"] = request.get("product_name") or ""
    ws["D5"] = request.get("spec") or ""
    ws["E5"] = request.get("quantity") or ""
    ws["F5"] = request.get("unit") or ""
    ws["B7"] = request.get("remark") or ""
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "14.3_\u51fa\u8ca8\u55ae.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_shipping_inspection_125(payload: dict):
    request = payload.get("shipment_request") or {}
    quality_rows = payload.get("quality_records") or []
    wb, ws = _new_wb("\u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304")
    ws["A1"] = "12.5 \u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304\u8868"
    ws["B3"] = request.get("order_no") or ""
    ws["C3"] = request.get("product_name") or ""
    ws["D3"] = request.get("spec") or ""
    ws["E3"] = request.get("quantity") or ""
    ws.append([])
    ws.append(["\u6750\u6599\u540d\u7a31", "\u6279\u865f", "\u7d50\u679c", "\u5099\u8a3b"])
    for row in quality_rows[:20]:
        ws.append([row.get("materialName"), row.get("batchNo"), row.get("result"), row.get("note")])
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "12.5_\u51fa\u8ca8\u6aa2\u67e5\u7d00\u9304\u8868.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_shipping_check_145(payload: dict):
    request = payload.get("shipment_request") or {}
    wb, ws = _new_wb("\u51fa\u8ca8\u6aa2\u67e5\u8868")
    ws["A1"] = "14.5 \u51fa\u8ca8\u6aa2\u67e5\u8868"
    ws["B3"] = request.get("order_no") or ""
    ws["C3"] = request.get("product_name") or ""
    ws["D3"] = request.get("batch_display") or request.get("order_no") or ""
    ws["B5"] = "\u5916\u89c0\u78ba\u8a8d"
    ws["C5"] = "\u5f85\u6aa2"
    ws["B6"] = "\u6a19\u793a\u78ba\u8a8d"
    ws["C6"] = "\u5f85\u6aa2"
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "14.5_\u51fa\u8ca8\u6aa2\u67e5\u8868.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_cip_tracker(payload: dict):
    item = _nonconformance_item(payload)
    wb, ws = _new_wb("CIP\u7d00\u9304")
    ws["A1"] = "15.2 \u88fd\u7a0b\u7f3a\u9677\u8ffd\u8e64\u6539\u5584(CIP)\u7d00\u9304\u8868"
    ws["A2"] = "\u6b04\u4f4d"
    ws["B2"] = "\u5167\u5bb9"
    labels = [
        ("\u4e0d\u7b26\u5408\u985e\u578b", item.get("type")),
        ("\u4e0d\u7b26\u5408\u7de8\u865f", item.get("id")),
        ("\u767c\u751f\u65e5", item.get("date")),
        ("\u90e8\u9580", item.get("dept")),
        ("\u56b4\u91cd\u5ea6", item.get("severity")),
        ("\u554f\u984c\u63cf\u8ff0", item.get("description")),
        ("\u539f\u56e0\u5206\u6790", item.get("rootCause")),
        ("\u6539\u5584\u63aa\u65bd", item.get("correctiveAction")),
        ("\u8cac\u4efb\u4eba", item.get("responsible")),
        ("\u5230\u671f\u65e5", item.get("dueDate")),
        ("\u72c0\u614b", item.get("status")),
        ("\u7d50\u6848\u65e5\u671f", item.get("closeDate")),
        ("\u6548\u679c\u78ba\u8a8d", item.get("effectiveness")),
        ("\u5efa\u8b70\u8ffd\u8e64\u52d5\u4f5c", _nonconformance_next_action(item)),
        ("\u4f86\u6e90\u6a94\u6848", item.get("source_file")),
    ]
    row_no = 3
    for label, value in labels:
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value or ""
        row_no += 1

    follow_ws = wb.create_sheet("\u6539\u5584\u8ffd\u8e64")
    follow_ws.append(["\u8ffd\u8e64\u9805\u76ee", "\u76ee\u524d\u5167\u5bb9", "\u5efa\u8b70"])
    follow_rows = [
        ("\u554f\u984c\u63cf\u8ff0", item.get("description"), "\u78ba\u8a8d\u554f\u984c\u73fe\u8c61\u662f\u5426\u5df2\u5b8c\u6574\u7d00\u9304"),
        ("\u539f\u56e0\u5206\u6790", item.get("rootCause"), "\u5efa\u8b70\u88dc\u4e0a 5 Why \u6216\u7570\u5e38\u539f\u56e0\u4f9d\u64da"),
        ("\u6539\u5584\u63aa\u65bd", item.get("correctiveAction"), "\u5efa\u8b70\u5206\u6210\u61c9\u6025\u63aa\u65bd\u8207\u6c38\u4e45\u5c0d\u7b56"),
        ("\u8cac\u4efb\u4eba", item.get("responsible"), "\u786c\u6027\u8cac\u4efb\u4eba\u4e0d\u5efa\u8b70\u7559\u7a7a"),
        ("\u6642\u9593\u63a7\u5236", item.get("dueDate"), "\u5efa\u8b70\u8a2d\u5b9a\u5b8c\u6210\u671f\u9650\u8207\u8ffd\u8e64\u7bc0\u9ede"),
        ("\u72c0\u614b / \u6548\u679c", f"{item.get('status')} / {item.get('effectiveness')}", "\u7d50\u6848\u524d\u5efa\u8b70\u88dc\u4e0a\u6548\u679c\u78ba\u8a8d"),
    ]
    for label, value, note in follow_rows:
        follow_ws.append([label, value or "", note])
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "15.2_CIP\u7d00\u9304\u8868.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_nonconformance_summary(payload: dict):
    item = _nonconformance_item(payload)
    wb, ws = _new_wb("\u4e0d\u7b26\u5408\u4f86\u6e90\u6458\u8981")
    ws["A1"] = "\u4e0d\u7b26\u5408\u4f86\u6e90\u6458\u8981"
    labels = [
        ("\u7de8\u865f", item.get("id")),
        ("\u65e5\u671f", item.get("date")),
        ("\u90e8\u9580", item.get("dept")),
        ("\u985e\u578b", item.get("type")),
        ("\u56b4\u91cd\u5ea6", item.get("severity")),
        ("\u554f\u984c\u63cf\u8ff0", item.get("description")),
        ("\u539f\u56e0\u5206\u6790", item.get("rootCause")),
        ("\u6539\u5584\u63aa\u65bd", item.get("correctiveAction")),
        ("\u8cac\u4efb\u4eba", item.get("responsible")),
        ("\u5230\u671f\u65e5", item.get("dueDate")),
        ("\u72c0\u614b", item.get("status")),
        ("\u7d50\u6848\u65e5\u671f", item.get("closeDate")),
        ("\u6548\u679c\u78ba\u8a8d", item.get("effectiveness")),
        ("\u4f86\u6e90\u6a94\u6848", item.get("source_file")),
    ]
    row_no = 3
    for label, value in labels:
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value or ""
        row_no += 1
    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, "\u4e0d\u7b26\u5408\u4f86\u6e90\u6458\u8981.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _zip_files(entries, name):
    path = _tmp_path('.zip')
    with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path, file_name, _mimetype in entries:
            archive.write(file_path, arcname=file_name)
    return path, name, 'application/zip'


def _build_shipment_pack(payload: dict):
    request = payload.get("shipment_request") or {}
    order_no = _normalize_text(request.get("order_no")) or 'shipment'
    entries = [
        _build_shipment_order_143(payload),
        _build_shipping_inspection_125(payload),
        _build_shipping_check_145(payload),
    ]
    return _zip_files(entries, f"{order_no}_\u51fa\u8ca8\u6d41\u7a0b\u7d00\u9304\u5305.zip")


def _build_cip_pack(payload: dict):
    item = _nonconformance_item(payload)
    record_id = _normalize_text(item.get("id")) or 'NC'
    entries = [
        _build_cip_tracker(payload),
        _build_nonconformance_summary(payload),
    ]
    return _zip_files(entries, f"{record_id}_CIP\u4e0d\u7b26\u5408\u6d41\u7a0b\u5305.zip")


def _build_env_monthly_pack(payload: dict):
    records = list(payload.get("env_records") or [])
    period_tag, _period_label = _env_pack_period(records)
    entries = [
        _build_env_monthly_summary(payload),
        _build_env_record(payload),
    ]
    return _zip_files(entries, f"{period_tag}_\u74b0\u5883\u76e3\u63a7\u6708\u5831\u5305.zip")


# ── MP 條文名稱對照表 ──────────────────────────────────────
_MP_NAMES: dict[str, str] = {
    "MP-01": "品質手冊管制", "MP-02": "文件管制", "MP-03": "記錄管制",
    "MP-04": "管理審查",     "MP-05": "量測資源管理", "MP-06": "工作環境管理",
    "MP-07": "人力資源",     "MP-08": "設計與開發", "MP-09": "內部稽核",
    "MP-10": "供應商管理",   "MP-11": "生產管制", "MP-12": "採購管制",
    "MP-13": "不合格品管制", "MP-14": "矯正措施", "MP-15": "持續改進",
}


def _audit_plan_item(payload: dict) -> dict:
    """從 payload 中找出選定的稽核計畫記錄（或第一筆）。"""
    plans = list(payload.get("audit_plans") or [])
    selected_id = _normalize_text(str(payload.get("selected_audit_id") or ""))
    if selected_id:
        for p in plans:
            if _normalize_text(str(p.get("id") or "")) == selected_id:
                return dict(p)
    return dict(plans[0]) if plans else {}


def _audit_required_details(payload: dict) -> list[dict]:
    """驗證稽核計畫必要欄位，回傳缺失明細清單。"""
    plans = list(payload.get("audit_plans") or [])
    if not plans:
        return [_missing_detail("audit_plans", "audit_plans", "稽核計畫資料", scope_label="稽核計畫")]
    item = _audit_plan_item(payload)
    if not item:
        return [_missing_detail("audit_plans", "audit_plans", "找不到選定的稽核計畫", scope_label="稽核計畫")]
    required = [
        ("scheduledDate", "預定日期"),
        ("dept", "受稽部門"),
        ("auditor", "稽核員"),
        ("auditee", "受稽代表"),
    ]
    return [
        _missing_detail("audit_plans", fk, label, scope_label="稽核計畫")
        for fk, label in required
        if not _normalize_text(str(item.get(fk) or ""))
    ]


def _build_audit_notice(payload: dict):
    """產生內部稽核通知書（含查檢範圍工作表）。"""
    item = _audit_plan_item(payload)
    audit_id = _normalize_text(str(item.get("id") or "")) or "IA"
    scope_str = _normalize_text(str(item.get("scope") or "")) or ""
    mp_codes = [c.strip() for c in scope_str.replace("，", ",").split(",") if c.strip()]

    wb, ws = _new_wb("內部稽核通知書")
    ws["A1"] = "潔沛企業有限公司"
    ws["A2"] = "9 內部稽核通知書"
    ws["A4"] = "稽核編號";    ws["B4"] = item.get("id") or ""
    ws["A5"] = "稽核類別";    ws["B5"] = "內部稽核（ISO 9001:2015）"
    ws["A6"] = "年度";        ws["B6"] = item.get("year") or ""
    ws["A7"] = "期別";        ws["B7"] = item.get("period") or ""
    ws["A8"] = "預定日期";    ws["B8"] = item.get("scheduledDate") or ""
    ws["A9"] = "實際日期";    ws["B9"] = item.get("actualDate") or "（待確定）"
    ws["A10"] = "受稽部門";   ws["B10"] = item.get("dept") or ""
    ws["A11"] = "稽核範圍";   ws["B11"] = scope_str
    ws["A12"] = "稽核員";     ws["B12"] = item.get("auditor") or ""
    ws["A13"] = "受稽代表";   ws["B13"] = item.get("auditee") or ""
    ws["A15"] = "備註";       ws["B15"] = "（稽核員填寫）"
    ws["A17"] = "主管核准";   ws["B17"] = ""

    # 查檢範圍工作表
    scope_ws = wb.create_sheet("稽核查檢範圍")
    scope_ws.append(["條文代碼", "條文名稱", "查驗完成（Y/N）", "主要發現 / 備註"])
    for code in mp_codes:
        scope_ws.append([code, _MP_NAMES.get(code, ""), "", ""])
    if not mp_codes:
        scope_ws.append(["（請填入條文代碼）", "", "", ""])

    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, f"{audit_id}_內部稽核通知書.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_audit_pack(payload: dict) -> tuple:
    """產生稽核流程包（通知書 ZIP）。"""
    item = _audit_plan_item(payload)
    audit_id = _normalize_text(str(item.get("id") or "")) or "IA"
    entries = [_build_audit_notice(payload)]
    return _zip_files(entries, f"{audit_id}_稽核流程包.zip")


def _build_management_review_summary(payload: dict):
    """產生管理審查輸入摘要工作簿（多工作表）。"""
    today_str = datetime.today().strftime("%Y-%m-%d")
    year_month = datetime.today().strftime("%Y%m")
    year = datetime.today().strftime("%Y")

    prod_records = list(payload.get("prod_records") or [])
    quality_records = list(payload.get("quality_records") or [])
    env_records = _sort_env_records(list(payload.get("env_records") or []))
    audit_plans = list(payload.get("audit_plans") or [])
    all_nc = list(payload.get("all_nonconformances") or [])

    # 預算統計
    nc_total = len(all_nc)
    nc_open = sum(1 for x in all_nc if _normalize_text(x.get("status")) not in ("已關閉", "Closed"))
    audit_done = sum(1 for x in audit_plans if _normalize_text(x.get("status")) == "已完成")
    audit_findings = sum(x.get("findings") or 0 for x in audit_plans)
    audit_nc = sum(x.get("ncCount") or 0 for x in audit_plans)
    env_total = len(env_records)
    env_fail = sum(1 for x in env_records if _normalize_text(x.get("result")) == "不合格")
    env_warn = sum(1 for x in env_records if _normalize_text(x.get("result")) == "警告")

    wb = Workbook()

    # ── 封面 ────────────────────────────────────────────
    cover = wb.active
    cover.title = "封面"
    cover["A1"] = "潔沛企業有限公司"
    cover["A1"].font = Font(bold=True, size=18, color="1E3A5F", name="Calibri")
    cover["A2"] = "16 管理審查輸入摘要報告"
    cover["A2"].font = Font(bold=True, size=14, color="2563EB", name="Calibri")
    cover["A3"] = f"審查年度：{year} 年"
    cover["A3"].font = Font(size=11, color="475569", name="Calibri")

    info_rows = [
        ("項目", "內容"),
        ("審查日期", today_str),
        ("主持人", ""),
        ("出席部門", "品管 / 生產 / 採購 / 業務 / 管理"),
        ("審查範圍", "QMS 全系統"),
        ("下次審查預定", ""),
    ]
    for r_idx, (label, val) in enumerate(info_rows, start=5):
        cover.cell(r_idx, 1, label)
        cover.cell(r_idx, 2, val)
    _style_header_row(cover, row=5, fill_hex="1E3A5F")
    for r_idx in range(6, 5 + len(info_rows)):
        for col in [1, 2]:
            cover.cell(r_idx, col).border = _BORDER

    # 資料來源摘要小表
    cover["A13"] = "資料來源摘要"
    cover["A13"].font = Font(bold=True, color="1E3A5F", name="Calibri")
    data_rows = [
        ("資料類別", "筆數", "說明"),
        ("不符合記錄", nc_total, f"未關閉 {nc_open} 筆" if nc_total else "無記錄"),
        ("內部稽核計畫", len(audit_plans), f"已完成 {audit_done} 次，發現 {audit_findings} 項" if audit_plans else "無記錄"),
        ("環境監控記錄", env_total, f"異常 {env_warn + env_fail} 筆" if env_total else "無記錄"),
        ("生產批次記錄", len(prod_records), ""),
        ("進料檢驗記錄", len(quality_records), ""),
    ]
    for r_idx, row_data in enumerate(data_rows, start=14):
        for col_idx, val in enumerate(row_data, start=1):
            cover.cell(r_idx, col_idx, val)
    _style_header_row(cover, row=14, fill_hex="1E4E8C")
    for r_idx in range(15, 14 + len(data_rows)):
        for col in range(1, 4):
            cover.cell(r_idx, col).border = _BORDER

    cover["A22"] = "製表人";    cover["B22"] = ""
    cover["A23"] = "主管核准";  cover["B23"] = ""
    cover["A24"] = "製表日期";  cover["B24"] = today_str
    for r in [22, 23, 24]:
        for col in [1, 2]:
            cover.cell(r, col).border = _BORDER
    cover.column_dimensions["A"].width = 18
    cover.column_dimensions["B"].width = 40
    cover.column_dimensions["C"].width = 36

    # ── 輸入項目清單（ISO 9001 9.3.2）──────────────────
    inputs_ws = wb.create_sheet("輸入項目清單")
    inputs_ws["A1"] = "管理審查輸入項目（依 ISO 9001:2015 第 9.3.2 條）"
    inputs_ws["A1"].font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    header = ["項次", "輸入項目", "資料來源", "本期摘要（自動帶入）", "評估 / 備註"]
    for col_idx, h in enumerate(header, start=1):
        inputs_ws.cell(3, col_idx, h)
    _style_header_row(inputs_ws, row=3, fill_hex="1E3A5F")

    inputs_data = [
        ("a",   "上次管理審查行動事項追蹤",     "前次決議記錄",           "（請填入上次決議追蹤狀況）"),
        ("b",   "外部與內部議題變更",             "環境分析、法規更新",     "（請填入本期重要變化）"),
        ("c-1", "顧客滿意度與回饋",               "客訴記錄、滿意度調查",   "（請填入客訴筆數與滿意度）"),
        ("c-2", "品質目標達成情形",               "生產記錄、檢驗報告",
         f"生產 {len(prod_records)} 批次，進料檢驗 {len(quality_records)} 筆"),
        ("c-3", "製程績效與產品符合性",           "生產日報、品質記錄",
         f"共 {len(prod_records)} 批次，詳見「品質績效摘要」工作表"),
        ("c-4", "不符合事項與矯正措施",           "不符合管理記錄",
         f"共 {nc_total} 筆，未關閉 {nc_open} 筆，詳見「不符合統計」"),
        ("c-5", "量測監控結果",                   "環境監控、儀器校正",
         f"環境監控 {env_total} 筆，異常 {env_warn + env_fail} 筆"),
        ("c-6", "稽核結果",                       "內部稽核計畫",
         f"已執行 {audit_done} 次，發現 {audit_findings} 項，不符合 {audit_nc} 項"),
        ("c-7", "外部供應商績效",                 "供應商評鑑記錄",         "（請填入供應商評鑑摘要）"),
        ("d",   "資源適當性",                     "人力、設備、設施",       "（請填入資源需求評估）"),
        ("e",   "風險與機會行動有效性",           "風險分析記錄",           "（請填入風險控制評估）"),
        ("f",   "改善機會",                       "前述各項綜合評估",       "（請填入本期改善建議）"),
    ]
    for r_idx, row_data in enumerate(inputs_data, start=4):
        for col_idx, val in enumerate(row_data, start=1):
            cell = inputs_ws.cell(r_idx, col_idx, val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        inputs_ws.row_dimensions[r_idx].height = 28
    inputs_ws.column_dimensions["A"].width = 6
    inputs_ws.column_dimensions["B"].width = 26
    inputs_ws.column_dimensions["C"].width = 22
    inputs_ws.column_dimensions["D"].width = 42
    inputs_ws.column_dimensions["E"].width = 20

    # ── 品質績效摘要 ─────────────────────────────────────
    perf_ws = wb.create_sheet("品質績效摘要")
    perf_ws["A1"] = "品質績效摘要"
    perf_ws["A1"].font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    perf_ws["A3"] = "一、生產批次績效"
    perf_ws["A3"].font = Font(bold=True, color="1E4E8C", name="Calibri")
    prod_headers = ["批號", "客戶", "產品", "投入", "良品", "不良", "良率", "日期"]
    for col_idx, h in enumerate(prod_headers, start=1):
        perf_ws.cell(5, col_idx, h)
    _style_header_row(perf_ws, row=5, fill_hex="1E4E8C")
    for r_idx, item in enumerate(prod_records, start=6):
        perf_ws.cell(r_idx, 1, item.get("lot"))
        perf_ws.cell(r_idx, 2, item.get("customer"))
        perf_ws.cell(r_idx, 3, item.get("product"))
        perf_ws.cell(r_idx, 4, item.get("input"))
        perf_ws.cell(r_idx, 5, item.get("good"))
        perf_ws.cell(r_idx, 6, item.get("defect"))
        perf_ws.cell(r_idx, 7, item.get("yieldRate"))
        perf_ws.cell(r_idx, 8, item.get("date"))
        for col in range(1, 9):
            perf_ws.cell(r_idx, col).border = _BORDER
    if not prod_records:
        perf_ws.cell(6, 1, "（本期無生產記錄）")

    q_start = 6 + len(prod_records) + 3
    perf_ws.cell(q_start, 1, "二、進料檢驗績效")
    perf_ws.cell(q_start, 1).font = Font(bold=True, color="1E4E8C", name="Calibri")
    q_hdr_row = q_start + 2
    q_headers = ["材料名稱", "批號", "數量", "規格", "PH", "比重", "結果", "備註"]
    for col_idx, h in enumerate(q_headers, start=1):
        perf_ws.cell(q_hdr_row, col_idx, h)
    _style_header_row(perf_ws, row=q_hdr_row, fill_hex="1E4E8C")
    for r_idx, item in enumerate(quality_records, start=q_hdr_row + 1):
        perf_ws.cell(r_idx, 1, item.get("materialName"))
        perf_ws.cell(r_idx, 2, item.get("batchNo"))
        perf_ws.cell(r_idx, 3, item.get("quantity"))
        perf_ws.cell(r_idx, 4, item.get("spec"))
        perf_ws.cell(r_idx, 5, item.get("ph"))
        perf_ws.cell(r_idx, 6, item.get("density"))
        perf_ws.cell(r_idx, 7, item.get("result"))
        perf_ws.cell(r_idx, 8, item.get("note"))
        for col in range(1, 9):
            perf_ws.cell(r_idx, col).border = _BORDER
    if not quality_records:
        perf_ws.cell(q_hdr_row + 1, 1, "（本期無進料檢驗記錄）")
    _autofit_columns(perf_ws)

    # ── 不符合統計 ───────────────────────────────────────
    nc_ws = wb.create_sheet("不符合統計")
    nc_ws["A1"] = "不符合事項統計"
    nc_ws["A1"].font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    nc_ws["A3"] = "統計摘要"
    nc_ws["A3"].font = Font(bold=True, color="1E4E8C", name="Calibri")
    kpi_nc = [
        ("項目", "數值"),
        ("總筆數", nc_total),
        ("已關閉", nc_total - nc_open),
        ("未關閉", nc_open),
        ("關閉率", f"{round((nc_total - nc_open) / nc_total * 100, 1)}%" if nc_total else "N/A"),
    ]
    for r_idx, (label, val) in enumerate(kpi_nc, start=4):
        nc_ws.cell(r_idx, 1, label)
        nc_ws.cell(r_idx, 2, val)
    _style_header_row(nc_ws, row=4, fill_hex="7F1D1D" if nc_open > 0 else "1E4E8C")
    for r_idx in range(5, 4 + len(kpi_nc)):
        for col in [1, 2]:
            nc_ws.cell(r_idx, col).border = _BORDER

    nc_ws.cell(11, 1, "不符合記錄明細")
    nc_ws.cell(11, 1).font = Font(bold=True, color="1E4E8C", name="Calibri")
    nc_headers = ["編號", "發生日期", "部門", "類型", "嚴重度", "問題描述", "狀態", "到期日", "有效性"]
    for col_idx, h in enumerate(nc_headers, start=1):
        nc_ws.cell(12, col_idx, h)
    _style_header_row(nc_ws, row=12, fill_hex="1E3A5F")
    for r_idx, item in enumerate(all_nc, start=13):
        vals = [
            item.get("id"), item.get("date"), item.get("dept"),
            item.get("type"), item.get("severity"), item.get("description"),
            item.get("status"), item.get("dueDate"), item.get("effectiveness"),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = nc_ws.cell(r_idx, col_idx, val)
            cell.border = _BORDER
        if _normalize_text(item.get("status")) not in ("已關閉", "Closed"):
            for col in range(1, 10):
                nc_ws.cell(r_idx, col).fill = _hdr_fill("FEE2E2")
    if not all_nc:
        nc_ws.cell(13, 1, "（本期無不符合記錄）")
    _autofit_columns(nc_ws)

    # ── 稽核結果摘要 ─────────────────────────────────────
    audit_ws = wb.create_sheet("稽核結果摘要")
    audit_ws["A1"] = "內部稽核結果摘要"
    audit_ws["A1"].font = Font(bold=True, size=12, color="1E3A5F", name="Calibri")
    a_headers = ["稽核編號", "年度", "期別", "預定日期", "實際日期", "受稽部門", "稽核員", "狀態", "發現筆數", "不符合數"]
    for col_idx, h in enumerate(a_headers, start=1):
        audit_ws.cell(3, col_idx, h)
    _style_header_row(audit_ws, row=3, fill_hex="1E3A5F")
    for r_idx, item in enumerate(audit_plans, start=4):
        vals = [
            item.get("id"), item.get("year"), item.get("period"),
            item.get("scheduledDate"), item.get("actualDate"), item.get("dept"),
            item.get("auditor"), item.get("status"),
            item.get("findings") or 0, item.get("ncCount") or 0,
        ]
        for col_idx, val in enumerate(vals, start=1):
            audit_ws.cell(r_idx, col_idx, val).border = _BORDER
    if not audit_plans:
        audit_ws.cell(4, 1, "（本期無稽核計畫記錄）")
    _autofit_columns(audit_ws)

    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, f"16_管理審查輸入摘要_{year_month}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_management_review_decisions(payload: dict):
    """產生管理審查決議記錄表（含預設議題的空白模板）。"""
    today_str = datetime.today().strftime("%Y-%m-%d")
    year_month = datetime.today().strftime("%Y%m")

    wb, ws = _new_wb("決議記錄")
    ws["A1"] = "潔沛企業有限公司"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F", name="Calibri")
    ws["A2"] = "16 管理審查決議記錄表"
    ws["A2"].font = Font(bold=True, size=12, color="2563EB", name="Calibri")
    ws["A3"] = f"審查日期：{today_str}　　主持人：　　　　　出席人員："
    ws["A3"].font = Font(size=10, color="475569", name="Calibri")

    headers = ["議題", "討論重點摘要", "決議事項", "負責部門 / 人", "完成期限", "追蹤狀態"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(5, col_idx, h)
    _style_header_row(ws, row=5, fill_hex="1E3A5F")

    default_topics = [
        "上次審查行動事項追蹤",
        "品質目標達成評估",
        "製程績效與產品符合性",
        "不符合事項與矯正措施追蹤",
        "顧客滿意度與回饋",
        "內部稽核結果",
        "供應商績效評估",
        "資源需求評估",
        "改善機會與後續行動",
    ]
    for r_idx, topic in enumerate(default_topics, start=6):
        ws.cell(r_idx, 1, topic)
        for col in range(1, 7):
            ws.cell(r_idx, col).border = _BORDER
            ws.cell(r_idx, col).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r_idx].height = 36

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    sign_row = 6 + len(default_topics) + 2
    ws.cell(sign_row, 1, "主持人簽名：")
    ws.cell(sign_row, 3, "記錄人簽名：")
    ws.cell(sign_row, 5, "核准：")

    path = _tmp_xlsx()
    wb.save(path)
    wb.close()
    return path, f"16_管理審查決議記錄_{year_month}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_management_review_pack(payload: dict):
    """產生管理審查流程包（輸入摘要 + 決議記錄 ZIP）。"""
    year_month = datetime.today().strftime("%Y%m")
    entries = [
        _build_management_review_summary(payload),
        _build_management_review_decisions(payload),
    ]
    return _zip_files(entries, f"{year_month}_管理審查流程包.zip")


def generate_template(payload: dict):
    template_code = _normalize_text(payload.get("template_code"))
    if not template_code:
        raise ValueError("template_code is required.")
    precheck = precheck_template(payload)
    if not precheck["ready"]:
        raise ValueError("\u4ecd\u6709\u5fc5\u8981\u6b04\u4f4d\u672a\u5099\u9f4a\uff1a" + "\u3001".join(precheck["missing_items"]))
    if template_code == "env_record":
        return _build_env_record(payload)
    if template_code == "env_monthly_pack":
        return _build_env_monthly_pack(payload)
    if template_code == "production_daily":
        return _build_production_daily(payload)
    if template_code == "quality_incoming":
        return _build_quality_incoming(payload)
    if template_code == "material_request_112":
        return _build_material_request(payload)
    if template_code == "shipment_order_143":
        return _build_shipment_order_143(payload)
    if template_code == "shipping_inspection_125":
        return _build_shipping_inspection_125(payload)
    if template_code == "shipping_check_145":
        return _build_shipping_check_145(payload)
    if template_code == "shipping_pack":
        return _build_shipment_pack(payload)
    if template_code == "cip_152":
        return _build_cip_tracker(payload)
    if template_code == "cip_pack":
        return _build_cip_pack(payload)
    if template_code == "audit_notice":
        return _build_audit_notice(payload)
    if template_code == "audit_pack":
        return _build_audit_pack(payload)
    if template_code == "management_review_pack":
        return _build_management_review_pack(payload)
    raise ValueError(f"Unsupported template_code: {template_code}")


def build_engine_payload_snapshot(payload: dict) -> str:
    snapshot = {
        "template_code": payload.get("template_code"),
        "prompt": payload.get("prompt"),
        "shipment_request": payload.get("shipment_request"),
        "nonconformance": payload.get("nonconformance"),
        "counts": {
            "env": len(payload.get("env_records") or []),
            "production": len(payload.get("prod_records") or []),
            "quality": len(payload.get("quality_records") or []),
        },
    }
    return json.dumps(snapshot, ensure_ascii=False)
