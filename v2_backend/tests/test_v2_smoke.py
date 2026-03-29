from __future__ import annotations

from io import BytesIO
import unittest
import zipfile

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete

from configure_v2_postgres import normalize_postgres_url
from v2_backend.app.config import settings
from v2_backend.app.db import session_scope
from v2_backend.app.main import app
from v2_backend.app import models


class V2SmokeTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._original_openrouter_api_key = settings.openrouter_api_key
        object.__setattr__(settings, 'openrouter_api_key', '')
        cls.client = TestClient(app)

    @classmethod
    def tearDownClass(cls):
        object.__setattr__(settings, 'openrouter_api_key', cls._original_openrouter_api_key)

    def setUp(self):
        with session_scope() as session:
            session.execute(delete(models.CompareCache))
            session.execute(delete(models.AuditCache))
            session.execute(delete(models.AuditLog))

    def test_health(self):
        response = self.client.get('/api/v2/health')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['data']['service'], 'auto-audit-v2')

    def test_prompt_resolve(self):
        response = self.client.get('/api/v2/prompts/runtime/resolve', params={'task_type': 'doc_audit'})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['data']['task_type'], 'doc_audit')

    def test_spc_analyze(self):
        response = self.client.post(
            '/api/v2/spc/analyze',
            json={
                'parameter_name': 'Thickness',
                'values': [10.1, 10.0, 9.9, 10.2, 10.1],
                'lsl': 9.5,
                'usl': 10.5,
                'target': 10.0,
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertIn('metrics', payload['data'])

    def test_knowledge_qa(self):
        ingest_response = self.client.post(
            '/api/v2/documents/ingest',
            json={'paths': ['1 文件化資訊管制程序/文件資訊與知識管制程序2.0.docx'], 'metadata': {}},
        )
        document_id = ingest_response.json()['data']['documents'][0]['document_id']
        response = self.client.post(
            '/api/v2/knowledge/qa',
            json={'question': '文件化資訊管制程序是否提到保存？', 'limit': 3, 'document_id': document_id},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertIn('answer', payload['data'])
        self.assertEqual(payload['data']['source_document_ids'], [document_id])

    def test_document_audit_cache_hit(self):
        audit_payload = {'path': '1 文件化資訊管制程序/文件資訊與知識管制程序2.0.docx'}
        first = self.client.post('/api/v2/documents/audit', json=audit_payload).json()
        second = self.client.post('/api/v2/documents/audit', json=audit_payload).json()
        self.assertTrue(first['success'])
        self.assertTrue(second['success'])
        self.assertFalse(first['data']['cache_hit'])
        self.assertTrue(second['data']['cache_hit'])

    def test_cache_status_and_clear(self):
        audit_payload = {'path': '1 文件化資訊管制程序/文件資訊與知識管制程序2.0.docx'}
        compare_payload = {
            'left_path': '9 內部稽核管理程序/內部稽核管理程序.docx',
            'right_path': '16 管理審查程序/管理審查程序.docx',
            'use_llm': False,
        }
        self.client.post('/api/v2/documents/audit', json=audit_payload)
        self.client.post('/api/v2/documents/compare', json=compare_payload)

        status_response = self.client.get('/api/v2/cache/status')
        self.assertEqual(status_response.status_code, 200)
        status_payload = status_response.json()
        self.assertTrue(status_payload['success'])
        self.assertGreaterEqual(status_payload['data']['audit_cache_count'], 1)
        self.assertGreaterEqual(status_payload['data']['compare_cache_count'], 1)

        clear_response = self.client.post('/api/v2/cache/clear', params={'target': 'all'})
        self.assertEqual(clear_response.status_code, 200)
        clear_payload = clear_response.json()
        self.assertTrue(clear_payload['success'])
        self.assertGreaterEqual(clear_payload['data']['deleted_audit_cache'], 1)
        self.assertGreaterEqual(clear_payload['data']['deleted_compare_cache'], 1)

    def test_document_compare(self):
        response = self.client.post(
            '/api/v2/documents/compare',
            json={
                'left_path': '9 內部稽核管理程序/內部稽核管理程序.docx',
                'right_path': '16 管理審查程序/管理審查程序.docx',
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertIn('summary', payload['data'])
        self.assertEqual(len(payload['data']['source_document_ids']), 2)
        self.assertNotIn('compare_llm_summary', payload['data']['tool_outputs_used'])

    def test_same_family_compare_has_version_conclusion(self):
        response = self.client.post(
            '/api/v2/documents/compare',
            json={
                'left_path': '0  品質手冊/公司品質手冊(1.0)0824.docx',
                'right_path': '0  品質手冊/公司品質手冊(2.0).docx',
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertTrue(payload['data']['same_document_family'])
        self.assertIn('版次', payload['data']['version_change_conclusion'])

    def test_compare_cache_hit(self):
        compare_payload = {
            'left_path': '9 內部稽核管理程序/內部稽核管理程序.docx',
            'right_path': '16 管理審查程序/管理審查程序.docx',
            'use_llm': False,
        }
        first = self.client.post('/api/v2/documents/compare', json=compare_payload).json()
        second = self.client.post('/api/v2/documents/compare', json=compare_payload).json()
        self.assertTrue(first['success'])
        self.assertTrue(second['success'])
        self.assertFalse(first['data']['cache_hit'])
        self.assertTrue(second['data']['cache_hit'])

    def test_version_candidates(self):
        response = self.client.post(
            '/api/v2/documents/version-candidates',
            json={'path': '0  品質手冊/公司品質手冊(2.0).docx', 'limit': 6},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        candidate_paths = [item['path'] for item in payload['data']['candidates']]
        self.assertTrue(any('公司品質手冊(1.0)0824.docx' in item for item in candidate_paths))

    def test_document_compare_export(self):
        response = self.client.post(
            '/api/v2/documents/compare/export',
            json={
                'left_path': '9 內部稽核管理程序/內部稽核管理程序.docx',
                'right_path': '16 管理審查程序/管理審查程序.docx',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers['content-type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn('摘要', workbook.sheetnames)
        self.assertIn('文字差異', workbook.sheetnames)
        self.assertEqual(workbook['摘要']['A1'].value, '文件差異比對報告')

    def test_document_compare_export_docx(self):
        response = self.client.post(
            '/api/v2/documents/compare/export/docx',
            json={
                'left_path': '9 內部稽核管理程序/內部稽核管理程序.docx',
                'right_path': '16 管理審查程序/管理審查程序.docx',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers['content-type'],
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )
        archive = zipfile.ZipFile(BytesIO(response.content))
        self.assertIn('word/document.xml', archive.namelist())
        xml_text = archive.read('word/document.xml').decode('utf-8', errors='ignore')
        self.assertIn('潔沛企業有限公司', xml_text)
        self.assertIn('ISO 文件差異與版次覆核報告', xml_text)

    def test_document_audit_export_docx(self):
        response = self.client.post(
            '/api/v2/documents/audit/export/docx',
            json={'path': '1 文件化資訊管制程序/文件資訊與知識管制程序2.0.docx'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers['content-type'],
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )
        archive = zipfile.ZipFile(BytesIO(response.content))
        self.assertIn('word/document.xml', archive.namelist())
        xml_text = archive.read('word/document.xml').decode('utf-8', errors='ignore')
        self.assertIn('AI 文件稽核正式報告', xml_text)
        self.assertIn('潔沛企業有限公司', xml_text)

    def test_history_runs(self):
        self.client.post(
            '/api/v2/documents/audit',
            json={'path': '1 文件化資訊管制程序/文件資訊與知識管制程序2.0.docx'},
        )
        self.client.post(
            '/api/v2/documents/compare',
            json={
                'left_path': '9 內部稽核管理程序/內部稽核管理程序.docx',
                'right_path': '16 管理審查程序/管理審查程序.docx',
            },
        )
        response = self.client.get('/api/v2/history/runs', params={'mode': 'all', 'limit': 10})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        task_types = [item['task_type'] for item in payload['data']['items']]
        self.assertIn('doc_audit', task_types)
        self.assertIn('doc_compare', task_types)

    def test_normalize_postgres_url(self):
        normalized = normalize_postgres_url('postgresql://user:pass@localhost:5432/auto_audit')
        self.assertTrue(normalized.startswith('postgresql+psycopg://'))


if __name__ == '__main__':
    unittest.main()
