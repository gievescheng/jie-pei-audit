from __future__ import annotations

import csv
from pathlib import Path

import httpx
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from .config import settings


def resolve_project_path(path_str: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        return path
    return (settings.project_root / path_str).resolve()


def parse_document(path_str: str) -> dict:
    path = resolve_project_path(path_str)
    suffix = path.suffix.lower()
    if not path.exists():
        raise FileNotFoundError(f"Document not found: {path_str}")

    if suffix == ".docx":
        return _parse_docx(path)
    if suffix == ".pdf":
        return _parse_pdf(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path)
    if suffix == ".csv":
        return _parse_csv(path)
    if suffix in {".txt", ".md"}:
        return _parse_text(path)
    raise ValueError(f"Unsupported document type: {suffix}")


def _chunk_lines(lines: list[str], *, page_no: int | None = None) -> list[dict]:
    chunks = []
    buffer = []
    current_len = 0
    for line in lines:
        cleaned = line.strip()
        if not cleaned:
            continue
        if current_len + len(cleaned) > 800 and buffer:
            text = "\n".join(buffer)
            chunks.append({"page_no": page_no, "section_name": text.splitlines()[0][:80], "content": text})
            buffer = []
            current_len = 0
        buffer.append(cleaned)
        current_len += len(cleaned)
    if buffer:
        text = "\n".join(buffer)
        chunks.append({"page_no": page_no, "section_name": text.splitlines()[0][:80], "content": text})
    return chunks


def _parse_docx(path: Path) -> dict:
    doc = DocxDocument(path)
    lines = [paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()]
    return {"title": path.stem, "file_type": "docx", "full_text": "\n".join(lines), "chunks": _chunk_lines(lines)}


def _parse_pdf(path: Path) -> dict:
    reader = PdfReader(str(path))
    chunks = []
    pages_text = []
    for idx, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            pages_text.append(text)
            chunks.extend(_chunk_lines(text.splitlines(), page_no=idx))
    return {"title": path.stem, "file_type": "pdf", "full_text": "\n\n".join(pages_text), "chunks": chunks}


def _parse_xlsx(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell not in (None, "")]
            if values:
                lines.append(" | ".join(values))
    return {"title": path.stem, "file_type": "xlsx", "full_text": "\n".join(lines), "chunks": _chunk_lines(lines)}


def _parse_csv(path: Path) -> dict:
    lines = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            values = [str(cell).strip() for cell in row if str(cell).strip()]
            if values:
                lines.append(" | ".join(values))
    return {"title": path.stem, "file_type": "csv", "full_text": "\n".join(lines), "chunks": _chunk_lines(lines)}


def _parse_text(path: Path) -> dict:
    text = path.read_text(encoding="utf-8-sig")
    return {"title": path.stem, "file_type": path.suffix.lower().lstrip("."), "full_text": text, "chunks": _chunk_lines(text.splitlines())}


# ── ERP-QMS bridge resolvers ─────────────────────────────────────────────────
#
# Strategy A (separate DBs / default):  HTTP GET to erp_qms_core API endpoint
# Strategy B (same PostgreSQL / dev):   Direct query via shared SQLAlchemy session
#
# All resolvers follow the same contract:
#   - return None  when the FK is None / empty
#   - return dict  {"id": ..., "resolved": True, ...} on success
#   - return dict  {"id": ..., "resolved": False, "error": "..."} on failure (never raise)


def resolve_user(
    auditor_id: str | None,
    *,
    session=None,
    erp_base_url: str | None = None,
) -> dict | None:
    """Resolve a User from erp_qms_core by UUID.

    Used to enrich AuditLog.auditor_id with full_name / role information.
    """
    if not auditor_id:
        return None
    if erp_base_url:
        try:
            r = httpx.get(f"{erp_base_url}/api/users/{auditor_id}", timeout=5)
            r.raise_for_status()
            return r.json()
        except Exception as exc:
            return {"id": auditor_id, "resolved": False, "error": str(exc)}
    if session:
        try:
            from erp_qms_core.backend.app.models import User  # type: ignore[import]
            entity = session.get(User, auditor_id)
            if entity:
                return {
                    "id": entity.id,
                    "resolved": True,
                    "full_name": getattr(entity, "full_name", ""),
                    "email": getattr(entity, "email", ""),
                }
        except Exception as exc:
            return {"id": auditor_id, "resolved": False, "error": str(exc)}
    return {"id": auditor_id, "resolved": False, "note": "configure ERP_BASE_URL or shared session"}


def resolve_department(
    owner_dept_id: str | None,
    *,
    session=None,
    erp_base_url: str | None = None,
) -> dict | None:
    """Resolve a Department from erp_qms_core by UUID.

    Used to enrich Document.owner_dept_id with department name / code.
    """
    if not owner_dept_id:
        return None
    if erp_base_url:
        try:
            r = httpx.get(f"{erp_base_url}/api/departments/{owner_dept_id}", timeout=5)
            r.raise_for_status()
            return r.json()
        except Exception as exc:
            return {"id": owner_dept_id, "resolved": False, "error": str(exc)}
    if session:
        try:
            from erp_qms_core.backend.app.models import Department  # type: ignore[import]
            entity = session.get(Department, owner_dept_id)
            if entity:
                return {
                    "id": entity.id,
                    "resolved": True,
                    "name": getattr(entity, "name", ""),
                    "code": getattr(entity, "code", ""),
                }
        except Exception as exc:
            return {"id": owner_dept_id, "resolved": False, "error": str(exc)}
    return {"id": owner_dept_id, "resolved": False, "note": "configure ERP_BASE_URL or shared session"}


def resolve_customer(
    customer_id: str | None,
    *,
    session=None,
    erp_base_url: str | None = None,
) -> dict | None:
    """Resolve a Customer from erp_qms_core by UUID.

    Used to enrich AuditCache.customer_id with name / contact information.
    """
    if not customer_id:
        return None
    if erp_base_url:
        try:
            r = httpx.get(f"{erp_base_url}/api/customers/{customer_id}", timeout=5)
            r.raise_for_status()
            return r.json()
        except Exception as exc:
            return {"id": customer_id, "resolved": False, "error": str(exc)}
    if session:
        try:
            from erp_qms_core.backend.app.models import Customer  # type: ignore[import]
            entity = session.get(Customer, customer_id)
            if entity:
                return {
                    "id": entity.id,
                    "resolved": True,
                    "name": getattr(entity, "name", ""),
                    "contact_person": getattr(entity, "contact_person", ""),
                }
        except Exception as exc:
            return {"id": customer_id, "resolved": False, "error": str(exc)}
    return {"id": customer_id, "resolved": False, "note": "configure ERP_BASE_URL or shared session"}


def resolve_supplier(
    supplier_id: str | None,
    *,
    session=None,
    erp_base_url: str | None = None,
) -> dict | None:
    """Resolve a Supplier from erp_qms_core by UUID.

    Used to enrich CompareCache.supplier_id with name / category information.
    """
    if not supplier_id:
        return None
    if erp_base_url:
        try:
            r = httpx.get(f"{erp_base_url}/api/suppliers/{supplier_id}", timeout=5)
            r.raise_for_status()
            return r.json()
        except Exception as exc:
            return {"id": supplier_id, "resolved": False, "error": str(exc)}
    if session:
        try:
            from erp_qms_core.backend.app.models import Supplier  # type: ignore[import]
            entity = session.get(Supplier, supplier_id)
            if entity:
                return {
                    "id": entity.id,
                    "resolved": True,
                    "name": getattr(entity, "name", ""),
                    "category": getattr(entity, "category", ""),
                }
        except Exception as exc:
            return {"id": supplier_id, "resolved": False, "error": str(exc)}
    return {"id": supplier_id, "resolved": False, "note": "configure ERP_BASE_URL or shared session"}


def maybe_call_openrouter(*, system_prompt: str, policy_prompt: str, user_prompt: str) -> str | None:
    if not settings.openrouter_api_key:
        return None
    combined_system = (system_prompt + "\n" + policy_prompt).strip()
    payload = {
        "model": settings.openrouter_model,
        "messages": [
            {"role": "system", "content": combined_system},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.2,
        "max_tokens": 900,
    }
    with httpx.Client(timeout=settings.openrouter_timeout) as client:
        response = client.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {settings.openrouter_api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
        response.raise_for_status()
        body = response.json()
        choices = body.get("choices") or []
        if not choices:
            return None
        return (((choices[0] or {}).get("message") or {}).get("content") or "").strip() or None


def maybe_call_ollama(*, system_prompt: str, policy_prompt: str, user_prompt: str,
                      model: str | None = None, timeout_seconds: int | None = None) -> str | None:
    if not settings.ollama_base_url:
        return None
    combined_system = (system_prompt + "\n" + policy_prompt).strip()
    payload = {
        "model": model or settings.ollama_model,
        "messages": [
            {"role": "system", "content": combined_system},
            {"role": "user", "content": user_prompt},
        ],
        "stream": False,
        "temperature": 0.2,
    }
    try:
        read_timeout = float(timeout_seconds or settings.ollama_timeout)
        timeout = httpx.Timeout(connect=10.0, read=read_timeout, write=30.0, pool=10.0)
        with httpx.Client(timeout=timeout) as client:
            response = client.post(
                f"{settings.ollama_base_url.rstrip('/')}/v1/chat/completions",
                headers={"Content-Type": "application/json"},
                json=payload,
            )
            response.raise_for_status()
            body = response.json()
            choices = body.get("choices") or []
            if not choices:
                return None
            msg = (choices[0] or {}).get("message") or {}
            content = (msg.get("content") or "").strip()
            if not content:
                content = (msg.get("reasoning") or "").strip()
            return content or None
    except Exception:
        return None


def call_llm(*, system_prompt: str, policy_prompt: str = "", user_prompt: str,
             task_type: str | None = None, model_override: str | None = None) -> str | None:
    """OpenRouter 優先，若未設定則依 task_type 路由到對應 Ollama 模型。"""
    from .config import resolve_model_for_task
    result = maybe_call_openrouter(system_prompt=system_prompt, policy_prompt=policy_prompt, user_prompt=user_prompt)
    if result:
        return result
    model, timeout = resolve_model_for_task(task_type, model_override)
    return maybe_call_ollama(system_prompt=system_prompt, policy_prompt=policy_prompt,
                             user_prompt=user_prompt, model=model, timeout_seconds=timeout)


def ollama_available() -> bool:
    """快速檢查 Ollama 是否可連線。"""
    try:
        with httpx.Client(timeout=3) as client:
            r = client.get(f"{settings.ollama_base_url.rstrip('/')}/api/tags")
            return r.status_code == 200
    except Exception:
        return False


def list_ollama_models() -> list[dict]:
    """查詢 Ollama 已安裝模型清單。"""
    try:
        with httpx.Client(timeout=5) as client:
            r = client.get(f"{settings.ollama_base_url.rstrip('/')}/api/tags")
            r.raise_for_status()
            return [
                {
                    "name": m.get("name", ""),
                    "size": m.get("size", 0),
                    "parameter_size": m.get("details", {}).get("parameter_size", ""),
                }
                for m in r.json().get("models", [])
            ]
    except Exception:
        return []
